from __future__ import annotations

import mimetypes
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import Body, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import and_, desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import IntegrityError

from ..access import get_user_whitelist_permissions
from ..access import is_super_admin
from ..auth import get_current_user, require_admin
from ..db import get_db
from ..magic_academy_schemas import (
    AdminReadingAudioStatisticsExportPayload,
    AudioMakeupPayload,
    AudioMakeupSettingPayload,
    MagicAudioUploadPayload,
)
from ..models import MagicAudioMakeupSetting, MagicAudioUpload, MagicReadingContent, MagicReadingContentTarget, MagicReadingSeries, User
from ..points_service import grant_points, record_reading_streak
from . import router
from ._utils import (
    AUDIO_EXTENSIONS,
    DEFAULT_AUDIO_MAKEUP_DAYS,
    MAX_AUDIO_SIZE,
    SOURCE_AUDIO_MAKEUP,
    SOURCE_AUDIO_USER_UPLOAD,
    SOURCE_WHITELIST_AUTO,
    _iso,
    _month_last_day,
    _now,
    _parse_month,
    _attachment_disposition,
    _safe_filename,
    _user_name,
    _expected_days,
    _xlsx_response,
)
from ._video_helpers import _ensure_auto_audio_checkin

AUTO_CHECKIN_PUBLIC_FILENAME = "录音打卡.m4a"
AUTO_CHECKIN_PUBLIC_SIZE = 221 * 1024
AUTO_CHECKIN_PUBLIC_MIME = "audio/x-m4a"
DEFAULT_READING_AUDIO_EXPORT_COLUMNS = [
    "reading_content_id",
    "reading_date",
    "push_time",
    "title",
    "target_summary",
    "employee_name",
    "department",
    "position",
    "should_complete",
    "is_completed",
    "uploaded_at",
    "is_makeup",
    "makeup_deadline",
    "current_status",
]
READING_CONTENT_STATUS_LABELS = {
    "active": "启用",
    "disabled": "已停用",
}


def _format_export_date(value: date | datetime | str | None) -> str:
    if not value:
        return "—"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime("%Y-%m-%d")
    except ValueError:
        return str(value)


def _format_export_time(value: datetime | str | None) -> str:
    if not value:
        return "—"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    text = str(value).strip()
    if not text:
        return "—"
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return text.replace("T", " ")[:19] if "T" in text else text


def _format_export_push_time(value: Any) -> str:
    if value is None:
        return "—"
    text = str(value).strip()
    if not text:
        return "—"
    if len(text) >= 5 and text[2] == ":":
        return text[:5]
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
        return parsed.strftime("%H:%M")
    except ValueError:
        return text


def _format_export_bool(value: Any) -> str:
    return "是" if bool(value) else "否"


def _format_export_text(value: Any) -> str:
    if value is None:
        return "—"
    text = str(value).strip()
    return text or "—"


def _format_export_rate(value: Any) -> str:
    try:
        rate = float(value or 0)
    except (TypeError, ValueError):
        rate = 0.0
    return f"{rate:.2f}%"


def _build_reading_audio_export_filename(month: str | None) -> str:
    month_text = (month or date.today().strftime("%Y-%m")).strip() or date.today().strftime("%Y-%m")
    return _safe_filename(f"读书打卡统计_{month_text}") + ".xlsx"


def _reading_target_matches_user(user: User, target: MagicReadingContentTarget) -> bool:
    target_type = (target.target_type or "").strip().lower()
    target_id = (target.target_id or "").strip()
    is_employee_role = user.role in {"user", "admin"}
    if target_type == "all":
        return is_employee_role
    if target_type == "all_newcomers":
        return is_employee_role and bool(user.is_newcomer)
    if target_type == "department":
        return (user.department or "").strip() == target_id
    if target_type == "position":
        return (user.position or "").strip() == target_id
    if target_type == "job_level":
        return (user.job_level or "M线").strip() == target_id
    if target_type == "employment_status":
        return (user.employment_status or "").strip() == target_id
    if target_type == "user":
        return str(user.id) == target_id
    return False


def _effective_push_at(content: MagicReadingContent) -> datetime:
    if content.push_at:
        return content.push_at
    return datetime.combine(content.reading_date, content.push_time or datetime.min.time())


def _reading_target_summary(targets: list[MagicReadingContentTarget]) -> str:
    if any((item.target_type or "").lower() == "all" for item in targets):
        return "全部员工"
    if any((item.target_type or "").lower() == "all_newcomers" for item in targets):
        return "仅新人"
    departments = [(item.target_id or "").strip() for item in targets if (item.target_type or "").lower() == "department" and (item.target_id or "").strip()]
    if departments:
        return f"部门：{'、'.join(departments)}"
    positions = [(item.target_id or "").strip() for item in targets if (item.target_type or "").lower() == "position" and (item.target_id or "").strip()]
    if positions:
        return f"岗位：{'、'.join(positions)}"
    job_levels = [(item.target_id or "").strip() for item in targets if (item.target_type or "").lower() == "job_level" and (item.target_id or "").strip()]
    if job_levels:
        return f"职级：{'、'.join(job_levels)}"
    employment_statuses = [(item.target_id or "").strip() for item in targets if (item.target_type or "").lower() == "employment_status" and (item.target_id or "").strip()]
    if employment_statuses:
        return f"在职状态：{'、'.join(employment_statuses)}"
    user_count = sum(1 for item in targets if (item.target_type or "").lower() == "user" and (item.target_id or "").strip())
    if user_count:
        return f"指定员工 {user_count} 人"
    return "未设置"


def _status_code_for_user_content(
    content: MagicReadingContent,
    upload: MagicAudioUpload | None,
    *,
    now: datetime,
) -> tuple[str, str]:
    if now < _effective_push_at(content):
        return "future", "未到推送时间"
    if upload is not None:
        return "completed", "已完成"
    if content.makeup_deadline_at and now > content.makeup_deadline_at:
        return "expired", "已过补卡截止时间"
    return "pending", "待完成"


async def _list_user_target_reading_contents(
    db: AsyncSession,
    *,
    user: User,
    target_date: date,
) -> list[MagicReadingContent]:
    result = await db.execute(
        select(MagicReadingContent)
        .where(
            MagicReadingContent.is_deleted.is_(False),
            MagicReadingContent.status == "active",
            MagicReadingContent.reading_date == target_date,
        )
        .order_by(MagicReadingContent.push_at.asc(), MagicReadingContent.id.asc())
    )
    contents = result.scalars().all()
    if not contents:
        return []
    content_ids = [int(item.id) for item in contents]
    target_rows = (
        await db.execute(
            select(MagicReadingContentTarget).where(MagicReadingContentTarget.content_id.in_(content_ids))
        )
    ).scalars().all()
    targets_map: dict[int, list[MagicReadingContentTarget]] = {}
    for item in target_rows:
        targets_map.setdefault(int(item.content_id), []).append(item)
    return [
        item
        for item in contents
        if any(_reading_target_matches_user(user, target) for target in targets_map.get(int(item.id), []))
    ]


async def _get_user_target_reading_content_or_404(
    db: AsyncSession,
    *,
    user: User,
    reading_content_id: int,
) -> MagicReadingContent:
    content = await db.get(MagicReadingContent, reading_content_id)
    if not content or content.is_deleted or (content.status or "").strip().lower() != "active":
        raise HTTPException(status_code=404, detail="读书内容不存在。")
    target_rows = (
        await db.execute(
            select(MagicReadingContentTarget).where(MagicReadingContentTarget.content_id == reading_content_id)
        )
    ).scalars().all()
    if not any(_reading_target_matches_user(user, target) for target in target_rows):
        raise HTTPException(status_code=403, detail="当前读书内容未推送给你。")
    return content


async def _assert_audio_submission_window(
    db: AsyncSession,
    *,
    user: User,
    reading_content_id: int | None = None,
    target_date: date | None = None,
    now: datetime,
) -> MagicReadingContent | None:
    if reading_content_id:
        content = await _get_user_target_reading_content_or_404(db, user=user, reading_content_id=reading_content_id)
        if now < _effective_push_at(content):
            raise HTTPException(status_code=400, detail="未到推送时间。")
        if content.makeup_deadline_at and now > content.makeup_deadline_at:
            raise HTTPException(status_code=400, detail="已超过补卡截止时间，无法补交。")
        return content
    if not target_date:
        return None
    contents = await _list_user_target_reading_contents(db, user=user, target_date=target_date)
    if not contents:
        return None
    visible_contents = [item for item in contents if _effective_push_at(item) <= now]
    if not visible_contents:
        raise HTTPException(status_code=400, detail="当前读书内容尚未到达推送时间，暂不能提交。")
    if all(item.makeup_deadline_at and now > item.makeup_deadline_at for item in visible_contents):
        raise HTTPException(status_code=400, detail="已超过补卡截止时间，无法补交。")
    return None


def _parse_optional_date(value: str | None, *, field_name: str) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} 格式应为 YYYY-MM-DD。") from exc


async def _build_admin_reading_targets_map(
    db: AsyncSession,
    content_ids: list[int],
) -> dict[int, list[MagicReadingContentTarget]]:
    if not content_ids:
        return {}
    result = await db.execute(
        select(MagicReadingContentTarget)
        .where(MagicReadingContentTarget.content_id.in_(content_ids))
        .order_by(MagicReadingContentTarget.id.asc())
    )
    mapping: dict[int, list[MagicReadingContentTarget]] = {}
    for item in result.scalars().all():
        mapping.setdefault(int(item.content_id), []).append(item)
    return mapping


async def _build_reading_series_map(
    db: AsyncSession,
    *,
    contents: list[MagicReadingContent],
) -> dict[int, MagicReadingSeries]:
    series_ids = sorted({int(item.series_id) for item in contents if item.series_id})
    if not series_ids:
        return {}
    result = await db.execute(
        select(MagicReadingSeries).where(MagicReadingSeries.id.in_(series_ids))
    )
    return {int(item.id): item for item in result.scalars().all()}


async def _build_user_map(
    db: AsyncSession,
    *,
    user_ids: list[int],
) -> dict[int, User]:
    if not user_ids:
        return {}
    result = await db.execute(select(User).where(User.id.in_(sorted(set(user_ids)))))
    return {int(item.id): item for item in result.scalars().all()}


def _content_matches_user(user: User, targets: list[MagicReadingContentTarget]) -> bool:
    return any(_reading_target_matches_user(user, target) for target in targets)


async def _list_filtered_reading_contents_for_admin(
    db: AsyncSession,
    *,
    admin: User,
    month: str | None,
    start_date: str | None,
    end_date: str | None,
    reading_content_id: int | None,
) -> list[MagicReadingContent]:
    stmt = select(MagicReadingContent).where(MagicReadingContent.is_deleted.is_(False))
    if not is_super_admin(admin):
        stmt = stmt.where(MagicReadingContent.created_by == admin.id)
    if reading_content_id:
        stmt = stmt.where(MagicReadingContent.id == reading_content_id)
    else:
        if start_date:
            stmt = stmt.where(MagicReadingContent.reading_date >= _parse_optional_date(start_date, field_name="start_date"))
        if end_date:
            stmt = stmt.where(MagicReadingContent.reading_date <= _parse_optional_date(end_date, field_name="end_date"))
        if month and not start_date and not end_date:
            month_start, month_end = _parse_month(month)
            stmt = stmt.where(MagicReadingContent.reading_date >= month_start, MagicReadingContent.reading_date <= month_end)
    stmt = stmt.order_by(MagicReadingContent.reading_date.asc(), MagicReadingContent.push_at.asc(), MagicReadingContent.id.asc())
    return (await db.execute(stmt)).scalars().all()


async def _list_filtered_employee_users(
    db: AsyncSession,
    *,
    department: str | None,
    user_id: int | None,
) -> list[User]:
    stmt = select(User).where(User.role.in_(["user", "admin"]), User.disabled.is_(False))
    if department:
        stmt = stmt.where(User.department == department)
    if user_id:
        stmt = stmt.where(User.id == user_id)
    stmt = stmt.order_by(User.id.asc())
    return (await db.execute(stmt)).scalars().all()


async def _build_content_expected_users_map(
    db: AsyncSession,
    *,
    contents: list[MagicReadingContent],
    users: list[User],
) -> dict[int, list[User]]:
    content_ids = [int(item.id) for item in contents]
    targets_map = await _build_admin_reading_targets_map(db, content_ids)
    expected_map: dict[int, list[User]] = {}
    for content in contents:
        targets = targets_map.get(int(content.id), [])
        expected_map[int(content.id)] = [user for user in users if _content_matches_user(user, targets)]
    return expected_map


async def _build_uploads_map_for_contents(
    db: AsyncSession,
    *,
    content_ids: list[int],
    user_ids: list[int],
) -> dict[tuple[int, int], MagicAudioUpload]:
    if not content_ids or not user_ids:
        return {}
    result = await db.execute(
        select(MagicAudioUpload)
        .where(
            MagicAudioUpload.reading_content_id.in_(content_ids),
            MagicAudioUpload.user_id.in_(user_ids),
            MagicAudioUpload.is_deleted.is_(False),
        )
        .order_by(MagicAudioUpload.uploaded_on.desc(), MagicAudioUpload.id.desc())
    )
    uploads_map: dict[tuple[int, int], MagicAudioUpload] = {}
    for item in result.scalars().all():
        key = (int(item.reading_content_id), int(item.user_id))
        uploads_map.setdefault(key, item)
    return uploads_map


async def _build_reading_content_user_rows(
    db: AsyncSession,
    *,
    admin: User,
    reading_content_id: int,
    department: str | None,
    user_id: int | None,
    status: str | None,
) -> tuple[MagicReadingContent, list[dict[str, Any]], str]:
    content = await db.get(MagicReadingContent, reading_content_id)
    if not content or content.is_deleted:
        raise HTTPException(status_code=404, detail="读书内容不存在。")
    if not is_super_admin(admin) and int(content.created_by) != int(admin.id):
        raise HTTPException(status_code=403, detail="无权查看该读书内容统计。")
    users = await _list_filtered_employee_users(db, department=department, user_id=user_id)
    expected_map = await _build_content_expected_users_map(db, contents=[content], users=users)
    expected_users = expected_map.get(int(content.id), [])
    uploads_map = await _build_uploads_map_for_contents(
        db,
        content_ids=[int(content.id)],
        user_ids=[int(item.id) for item in expected_users],
    )
    now = _now()
    rows: list[dict[str, Any]] = []
    for target in expected_users:
        upload = uploads_map.get((int(content.id), int(target.id)))
        status_code, status_text = _status_code_for_user_content(content, upload, now=now)
        row = {
            "reading_content_id": int(content.id),
            "user_id": int(target.id),
            "user_name": _user_name(target),
            "department_name": target.department or "",
            "position": target.position or "",
            "should_complete": True,
            "completed": upload is not None,
            "upload_id": int(upload.id) if upload else None,
            "uploaded_at": _iso(upload.uploaded_on) if upload else None,
            "is_makeup": bool(upload and (upload.source or "") == SOURCE_AUDIO_MAKEUP),
            "makeup_at": _iso(upload.uploaded_on) if upload and (upload.source or "") == SOURCE_AUDIO_MAKEUP else None,
            "remark": (upload.remark or "") if upload else "",
            "status": status_code,
            "status_text": status_text,
        }
        if status and status != "all" and status_code != status:
            continue
        rows.append(row)
    legacy_count_result = await db.execute(
        select(func.count(MagicAudioUpload.id)).where(
            MagicAudioUpload.reading_content_id.is_(None),
            MagicAudioUpload.is_deleted.is_(False),
            MagicAudioUpload.uploaded_date == content.reading_date,
        )
    )
    legacy_hint = ""
    if int(legacy_count_result.scalar_one() or 0) > 0:
        legacy_hint = "存在历史未绑定录音，未计入具体读书内容完成率。"
    return content, rows, legacy_hint


def _serialize_audio_record(
    item: MagicAudioUpload,
    user_map: dict[int, User] | None = None,
    *,
    reveal_whitelist: bool = True,
) -> dict[str, Any]:
    owner = user_map.get(item.user_id) if user_map else None
    source = item.source or SOURCE_AUDIO_USER_UPLOAD
    source_label = (
        "补卡" if source == SOURCE_AUDIO_MAKEUP
        else "白名单自动打卡" if source == SOURCE_WHITELIST_AUTO and reveal_whitelist
        else "用户上传" if source == SOURCE_WHITELIST_AUTO
        else "用户上传"
    )
    file_name = item.file_name or ""
    file_size = int(item.file_size or 0)
    mime_type = item.mime_type or ""
    remark = item.remark or ""
    if source == SOURCE_WHITELIST_AUTO and not reveal_whitelist:
        file_name = AUTO_CHECKIN_PUBLIC_FILENAME
        file_size = AUTO_CHECKIN_PUBLIC_SIZE
        mime_type = AUTO_CHECKIN_PUBLIC_MIME
        remark = ""
    return {
        "id": item.id,
        "user_id": item.user_id,
        "reading_content_id": int(item.reading_content_id) if item.reading_content_id else None,
        "user_name": _user_name(owner) if owner else "",
        "department": (owner.department or "") if owner else "",
        "file_name": file_name,
        "file_size": file_size,
        "file_type": mime_type,
        "remark": remark,
        "uploaded_date": _iso(item.uploaded_date),
        "uploaded_time": _iso(item.uploaded_on),
        "status": "已上传",
        "source": source,
        "source_label": source_label,
        "is_makeup": source == SOURCE_AUDIO_MAKEUP,
        "auto_checkin_by_whitelist": bool(item.auto_checkin_by_whitelist),
    }


def _reading_audio_export_columns() -> dict[str, dict[str, Any]]:
    return {
        "reading_content_id": {
            "title": "读书内容ID",
            "getter": lambda ctx: ctx["content_id"],
        },
        "series_name": {
            "title": "所属系列",
            "getter": lambda ctx: _format_export_text(ctx["series_name"]),
        },
        "reading_date": {
            "title": "读书日期",
            "getter": lambda ctx: _format_export_date(ctx["content"].reading_date),
        },
        "push_time": {
            "title": "推送时间",
            "getter": lambda ctx: _format_export_push_time(ctx["content_row"]["push_time"]),
        },
        "title": {
            "title": "标题",
            "getter": lambda ctx: _format_export_text(ctx["content"].title),
        },
        "target_summary": {
            "title": "推送对象",
            "getter": lambda ctx: _format_export_text(ctx["content_row"]["target_summary"]),
        },
        "pushed_count": {
            "title": "推送人数",
            "getter": lambda ctx: ctx["content_row"]["expected_count"],
        },
        "completion_rate": {
            "title": "完成率",
            "getter": lambda ctx: _format_export_rate(ctx["content_row"]["completion_rate"]),
        },
        "makeup_deadline": {
            "title": "补卡截止时间",
            "getter": lambda ctx: _format_export_time(ctx["content"].makeup_deadline_at),
        },
        "content_status": {
            "title": "内容状态",
            "getter": lambda ctx: READING_CONTENT_STATUS_LABELS.get((ctx["content"].status or "").strip().lower(), "—"),
        },
        "created_by_name": {
            "title": "创建人",
            "getter": lambda ctx: _format_export_text(ctx["created_by_name"]),
        },
        "created_at": {
            "title": "创建时间",
            "getter": lambda ctx: _format_export_time(ctx["content"].created_at),
        },
        "employee_id": {
            "title": "员工ID",
            "getter": lambda ctx: ctx["detail_row"]["user_id"],
        },
        "employee_name": {
            "title": "员工姓名",
            "getter": lambda ctx: _format_export_text(ctx["detail_row"]["user_name"]),
        },
        "department": {
            "title": "部门",
            "getter": lambda ctx: _format_export_text(ctx["detail_row"]["department_name"]),
        },
        "position": {
            "title": "岗位",
            "getter": lambda ctx: _format_export_text(ctx["detail_row"]["position"]),
        },
        "should_complete": {
            "title": "是否应完成",
            "getter": lambda ctx: _format_export_bool(ctx["detail_row"]["should_complete"]),
        },
        "is_completed": {
            "title": "是否完成",
            "getter": lambda ctx: _format_export_bool(ctx["detail_row"]["completed"]),
        },
        "uploaded_at": {
            "title": "上传时间",
            "getter": lambda ctx: _format_export_time(ctx["detail_row"]["uploaded_at"]),
        },
        "is_makeup": {
            "title": "是否补卡",
            "getter": lambda ctx: _format_export_bool(ctx["detail_row"]["is_makeup"]),
        },
        "current_status": {
            "title": "当前状态",
            "getter": lambda ctx: _format_export_text(ctx["detail_row"]["status_text"]),
        },
    }


def _resolve_reading_audio_export_columns(columns: list[str] | None) -> list[str]:
    requested = [str(item or "").strip() for item in (columns or []) if str(item or "").strip()]
    if not requested:
        return list(DEFAULT_READING_AUDIO_EXPORT_COLUMNS)
    allowed = _reading_audio_export_columns()
    invalid = [item for item in requested if item not in allowed]
    if invalid:
        raise HTTPException(status_code=400, detail=f"存在不支持的导出字段：{', '.join(invalid)}")
    return requested


async def _collect_admin_reading_audio_export_rows(
    db: AsyncSession,
    *,
    admin: User,
    month: str | None,
    start_date: str | None,
    end_date: str | None,
    reading_content_id: int | None,
    department: str | None,
    user_id: int | None,
    status: str | None,
) -> list[dict[str, Any]]:
    contents = await _list_filtered_reading_contents_for_admin(
        db,
        admin=admin,
        month=month,
        start_date=start_date,
        end_date=end_date,
        reading_content_id=reading_content_id,
    )
    if not contents:
        return []
    content_ids = [int(item.id) for item in contents]
    targets_map = await _build_admin_reading_targets_map(db, content_ids)
    users = await _list_filtered_employee_users(db, department=department, user_id=user_id)
    expected_map = await _build_content_expected_users_map(db, contents=contents, users=users)
    uploads_map = await _build_uploads_map_for_contents(
        db,
        content_ids=content_ids,
        user_ids=[int(item.id) for item in users],
    )
    series_map = await _build_reading_series_map(db, contents=contents)
    creator_map = await _build_user_map(db, user_ids=[int(item.created_by) for item in contents if item.created_by])
    now = _now()

    export_rows: list[dict[str, Any]] = []
    for content in contents:
        content_id = int(content.id)
        targets = targets_map.get(content_id, [])
        expected_users = expected_map.get(content_id, [])
        completed_count = 0
        status_codes: set[str] = set()
        detail_rows: list[dict[str, Any]] = []
        for target in expected_users:
            upload = uploads_map.get((content_id, int(target.id)))
            if upload is not None:
                completed_count += 1
            status_code, status_text = _status_code_for_user_content(content, upload, now=now)
            status_codes.add(status_code)
            detail_rows.append({
                "reading_content_id": content_id,
                "user_id": int(target.id),
                "user_name": _user_name(target),
                "department_name": target.department or "",
                "position": target.position or "",
                "should_complete": True,
                "completed": upload is not None,
                "upload_id": int(upload.id) if upload else None,
                "uploaded_at": _iso(upload.uploaded_on) if upload else None,
                "is_makeup": bool(upload and (upload.source or "") == SOURCE_AUDIO_MAKEUP),
                "status": status_code,
                "status_text": status_text,
            })
        if status and status != "all" and status not in status_codes:
            continue
        expected_count = len(expected_users)
        content_row = {
            "reading_content_id": content_id,
            "reading_date": _iso(content.reading_date),
            "push_at": _iso(_effective_push_at(content)),
            "push_time": content.push_time.isoformat() if content.push_time else _effective_push_at(content).time().isoformat(),
            "title": content.title,
            "target_summary": _reading_target_summary(targets),
            "expected_count": expected_count,
            "completed_count": completed_count,
            "pending_count": max(expected_count - completed_count, 0),
            "completion_rate": round((completed_count / expected_count) * 100, 2) if expected_count else 0,
            "makeup_deadline_at": _iso(content.makeup_deadline_at),
            "is_deadline_passed": bool(content.makeup_deadline_at and now > content.makeup_deadline_at),
            "has_checkins": completed_count > 0,
        }
        series = series_map.get(int(content.series_id)) if content.series_id else None
        creator = creator_map.get(int(content.created_by)) if content.created_by else None
        for detail_row in detail_rows:
            if status and status != "all" and detail_row["status"] != status:
                continue
            export_rows.append({
                "content": content,
                "content_id": content_id,
                "content_row": content_row,
                "detail_row": detail_row,
                "series_name": series.title if series else "",
                "created_by_name": _user_name(creator) if creator else "",
            })
    return export_rows


def _build_reading_audio_export_workbook(
    *,
    filename: str,
    columns: list[str],
    row_contexts: list[dict[str, Any]],
) -> StreamingResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="未安装 openpyxl，暂时无法导出 Excel。") from exc

    column_defs = _reading_audio_export_columns()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "打卡明细"
    headers = [column_defs[column]["title"] for column in columns]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row_context in row_contexts:
        sheet.append([column_defs[column]["getter"](row_context) for column in columns])
    sheet.freeze_panes = "A2"
    for index, column in enumerate(columns, start=1):
        title = column_defs[column]["title"]
        max_len = len(str(title))
        for row_index in range(2, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=index).value
            max_len = max(max_len, len(str(value or "")))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 10), 35)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = _attachment_disposition(filename)
    return response


def _build_audio_calendar_payload(
    month_start: date,
    month_last_day_value: date,
    uploads: list[MagicAudioUpload],
    user_map: dict[int, User] | None = None,
    aggregate_users: bool = False,
    reveal_whitelist: bool = True,
) -> list[dict[str, Any]]:
    grouped: dict[str, list[MagicAudioUpload]] = {}
    for item in uploads:
        key = item.uploaded_date.isoformat() if item.uploaded_date else None
        if not key:
            continue
        grouped.setdefault(key, []).append(item)
    today = date.today().isoformat()
    days: list[dict[str, Any]] = []
    cursor = month_start
    while cursor <= month_last_day_value:
        key = cursor.isoformat()
        items = grouped.get(key, [])
        uploaded_users = sorted({item.user_id for item in items})
        days.append({
            "date": key,
            "is_today": key == today,
            "is_future": key > today,
            "uploaded": bool(items),
            "count": len(items),
            "uploaded_user_count": len(uploaded_users),
            "records": [_serialize_audio_record(item, user_map, reveal_whitelist=reveal_whitelist) for item in items],
            "user_ids": uploaded_users if aggregate_users else [],
        })
        cursor += timedelta(days=1)
    return days


def _serialize_audio_makeup_setting(row: MagicAudioMakeupSetting | None) -> dict[str, Any]:
    return {
        "enabled": bool(row.enabled) if row else False,
        "make_up_days": int(row.make_up_days or 0) if row else DEFAULT_AUDIO_MAKEUP_DAYS,
        "audio_random_window_minutes": int(row.audio_random_window_minutes or 0) if row else 0,
        "video_random_window_minutes": int(row.video_random_window_minutes or 0) if row else 0,
        "description": (
            f"仅允许补最近 {int(row.make_up_days or 0)} 天内未完成的读书打卡"
            if row and bool(row.enabled) and int(row.make_up_days or 0) > 0
            else "当前未开启补卡"
        ),
    }


async def _get_audio_makeup_setting(
    db: AsyncSession,
    *,
    create: bool = False,
) -> MagicAudioMakeupSetting | None:
    result = await db.execute(
        select(MagicAudioMakeupSetting).order_by(MagicAudioMakeupSetting.id.asc()).limit(1)
    )
    row = result.scalar_one_or_none()
    if row or not create:
        return row
    row = MagicAudioMakeupSetting(enabled=False, make_up_days=DEFAULT_AUDIO_MAKEUP_DAYS)
    db.add(row)
    await db.flush()
    await db.refresh(row)
    return row


async def _has_audio_checkin_on_date(
    db: AsyncSession,
    user_id: int,
    target_date: date,
) -> bool:
    result = await db.execute(
        select(MagicAudioUpload.id).where(
            MagicAudioUpload.user_id == user_id,
            MagicAudioUpload.is_deleted.is_(False),
            MagicAudioUpload.uploaded_date == target_date,
        )
    )
    return result.scalar_one_or_none() is not None


async def _has_audio_checkin_on_content(
    db: AsyncSession,
    user_id: int,
    reading_content_id: int,
) -> bool:
    result = await db.execute(
        select(MagicAudioUpload.id).where(
            MagicAudioUpload.user_id == user_id,
            MagicAudioUpload.reading_content_id == reading_content_id,
            MagicAudioUpload.is_deleted.is_(False),
        )
    )
    return result.scalar_one_or_none() is not None


def _evaluate_audio_makeup_date(
    target_date: date,
    *,
    today: date,
    setting: MagicAudioMakeupSetting | None,
    has_record: bool,
) -> tuple[bool, str]:
    if target_date > today:
        return False, "不能补未来日期。"
    if target_date == today:
        return False, "今日打卡请直接走正常打卡流程。"
    if not setting or not setting.enabled or int(setting.make_up_days or 0) <= 0:
        return False, "当前未开启补卡。"
    if has_record:
        return False, "该日期已完成打卡。"
    delta_days = (today - target_date).days
    if delta_days <= 0:
        return False, "不能补未来日期。"
    if delta_days > int(setting.make_up_days or 0):
        return False, "补卡时间已过期。"
    return True, ""


@router.get("/my/audios")
async def list_my_audios(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    whitelist_permissions = await get_user_whitelist_permissions(db, user.id)
    await _ensure_auto_audio_checkin(db, user, whitelist_permissions)
    result = await db.execute(
        select(MagicAudioUpload)
        .where(MagicAudioUpload.user_id == user.id, MagicAudioUpload.is_deleted.is_(False))
        .order_by(desc(MagicAudioUpload.uploaded_on))
    )
    return [_serialize_audio_record(item) for item in result.scalars().all()]


@router.get("/admin/audio-makeup-setting")
async def get_audio_makeup_setting(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    del admin
    row = await _get_audio_makeup_setting(db)
    return _serialize_audio_makeup_setting(row)


@router.put("/admin/audio-makeup-setting")
async def update_audio_makeup_setting(
    payload: AudioMakeupSettingPayload,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_audio_makeup_setting(db, create=True)
    if not row:
        raise HTTPException(status_code=500, detail="补卡设置初始化失败。")
    row.enabled = payload.enabled
    row.make_up_days = int(payload.make_up_days or 0)
    row.audio_random_window_minutes = int(payload.audio_random_window_minutes or 0)
    row.video_random_window_minutes = int(payload.video_random_window_minutes or 0)
    row.updated_by = admin.id
    await db.flush()
    await db.refresh(row)
    return _serialize_audio_makeup_setting(row)


@router.get("/my/audios/makeup-options")
async def get_my_audio_makeup_options(
    month: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    whitelist_permissions = await get_user_whitelist_permissions(db, user.id)
    await _ensure_auto_audio_checkin(db, user, whitelist_permissions)
    setting = await _get_audio_makeup_setting(db)
    month_start, _ = _parse_month(month)
    month_last_day_value = _month_last_day(month_start)
    result = await db.execute(
        select(MagicReadingContent)
        .where(
            MagicReadingContent.is_deleted.is_(False),
            MagicReadingContent.status == "active",
            MagicReadingContent.reading_date >= month_start,
            MagicReadingContent.reading_date <= month_last_day_value,
        )
        .order_by(MagicReadingContent.reading_date.asc(), MagicReadingContent.push_at.asc(), MagicReadingContent.id.asc())
    )
    contents = result.scalars().all()
    content_ids = [int(item.id) for item in contents]
    target_rows = (
        await db.execute(
            select(MagicReadingContentTarget).where(MagicReadingContentTarget.content_id.in_(content_ids))
        )
    ).scalars().all() if content_ids else []
    targets_map: dict[int, list[MagicReadingContentTarget]] = {}
    for item in target_rows:
        targets_map.setdefault(int(item.content_id), []).append(item)
    visible_contents = [
        item for item in contents
        if any(_reading_target_matches_user(user, target) for target in targets_map.get(int(item.id), []))
    ]
    upload_stmt = (
        select(MagicAudioUpload)
        .where(
            MagicAudioUpload.user_id == user.id,
            MagicAudioUpload.is_deleted.is_(False),
            MagicAudioUpload.uploaded_date >= month_start,
            MagicAudioUpload.uploaded_date <= month_last_day_value,
        )
        .order_by(MagicAudioUpload.uploaded_on.asc())
    )
    if visible_contents:
        upload_stmt = upload_stmt.where(
            MagicAudioUpload.reading_content_id.in_([int(item.id) for item in visible_contents])
        )
    upload_result = await db.execute(upload_stmt)
    uploads = upload_result.scalars().all()
    uploads_by_content = {int(item.reading_content_id): item for item in uploads if item.reading_content_id}
    today = date.today()
    items = []
    for content in visible_contents:
        has_record = int(content.id) in uploads_by_content
        can_makeup, reason = _evaluate_audio_makeup_date(
            content.reading_date,
            today=today,
            setting=setting,
            has_record=has_record,
        )
        now = _now()
        if now < _effective_push_at(content):
            can_makeup = False
            reason = "未到推送时间。"
        elif content.makeup_deadline_at and now > content.makeup_deadline_at:
            can_makeup = False
            reason = "已超过补卡截止时间，无法补交。"
        items.append({
            "reading_content_id": int(content.id),
            "date": content.reading_date.isoformat(),
            "title": content.title,
            "push_at": _iso(_effective_push_at(content)),
            "makeup_deadline_at": _iso(content.makeup_deadline_at),
            "can_makeup": can_makeup,
            "reason": reason,
            "has_record": has_record,
            "is_future": content.reading_date > today,
            "is_expired": bool(reason in {"补卡时间已过期。", "已超过补卡截止时间，无法补交。"}),
        })
    return {
        "month": month_start.strftime("%Y-%m"),
        "setting": _serialize_audio_makeup_setting(setting),
        "days": items,
    }


@router.post("/my/audios")
async def upload_my_audio(
    payload: MagicAudioUploadPayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    suffix = Path(payload.file_name or "").suffix.lower()
    if suffix not in AUDIO_EXTENSIONS:
        raise HTTPException(status_code=400, detail="音频格式不支持。")
    if int(payload.file_size or 0) > MAX_AUDIO_SIZE:
        raise HTTPException(status_code=400, detail="单个录音文件不能超过 50MB。")
    safe_name = _safe_filename(payload.file_name or f"audio{suffix}")
    now = _now()
    content = await _assert_audio_submission_window(
        db,
        user=user,
        reading_content_id=int(payload.reading_content_id),
        target_date=now.date(),
        now=now,
    )
    if await _has_audio_checkin_on_content(db, user.id, int(payload.reading_content_id)):
        raise HTTPException(status_code=400, detail="该读书内容已完成，无需重复提交。")
    row = MagicAudioUpload(
        user_id=user.id,
        reading_content_id=int(payload.reading_content_id),
        file_name=safe_name,
        file_path="",
        file_size=int(payload.file_size or 0),
        mime_type=(payload.mime_type or mimetypes.guess_type(safe_name)[0] or suffix.lstrip(".")).strip(),
        remark=(payload.remark or "").strip(),
        source=SOURCE_AUDIO_USER_UPLOAD,
        auto_checkin_by_whitelist=False,
        uploaded_on=now,
        uploaded_date=content.reading_date if content else now.date(),
        is_deleted=False,
    )
    db.add(row)
    try:
        await db.flush()
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(status_code=400, detail="该读书内容已完成打卡，请勿重复提交。") from exc
    # 读书打卡积分（每日上限 1 次，dedupe_extra=日期 保证每日唯一）+ streak 奖励
    try:
        checkin_date = row.uploaded_date if isinstance(row.uploaded_date, date) else now.date()
        await grant_points(
            db,
            user_id=user.id,
            rule_code="reading_checkin",
            business_type="audio_upload",
            business_id=int(row.id),
            dedupe_extra=f"d{checkin_date.isoformat()}",
            remark=f"读书打卡 {checkin_date.isoformat()}",
        )
        await record_reading_streak(db, user_id=user.id, checkin_date=checkin_date)
    except Exception:  # noqa: BLE001
        pass
    return {
        "id": row.id,
        "file_name": row.file_name,
        "file_size": int(row.file_size or 0),
        "file_type": row.mime_type,
        "remark": row.remark or "",
        "uploaded_date": _iso(row.uploaded_date),
        "uploaded_time": _iso(row.uploaded_on),
        "status": "已上传",
        "source": SOURCE_AUDIO_USER_UPLOAD,
        "source_label": "用户上传",
    }


@router.post("/my/audios/makeup")
async def submit_my_audio_makeup(
    payload: AudioMakeupPayload,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    suffix = Path(payload.file_name or "").suffix.lower()
    if suffix not in AUDIO_EXTENSIONS:
        raise HTTPException(status_code=400, detail="音频格式不支持。")
    if int(payload.file_size or 0) > MAX_AUDIO_SIZE:
        raise HTTPException(status_code=400, detail="单个录音文件不能超过 50MB。")
    whitelist_permissions = await get_user_whitelist_permissions(db, user.id)
    await _ensure_auto_audio_checkin(db, user, whitelist_permissions)
    today = date.today()
    content = await _get_user_target_reading_content_or_404(db, user=user, reading_content_id=int(payload.reading_content_id))
    target_date = content.reading_date
    now = _now()
    setting = await _get_audio_makeup_setting(db)
    has_record = await _has_audio_checkin_on_content(db, user.id, int(payload.reading_content_id))
    can_makeup, reason = _evaluate_audio_makeup_date(
        target_date,
        today=today,
        setting=setting,
        has_record=has_record,
    )
    if not can_makeup:
        raise HTTPException(status_code=400, detail=reason)
    await _assert_audio_submission_window(
        db,
        user=user,
        reading_content_id=int(payload.reading_content_id),
        target_date=target_date,
        now=now,
    )
    safe_name = _safe_filename(payload.file_name or f"audio{suffix}")
    row = MagicAudioUpload(
        user_id=user.id,
        reading_content_id=int(payload.reading_content_id),
        file_name=safe_name,
        file_path="",
        file_size=int(payload.file_size or 0),
        mime_type=(payload.mime_type or mimetypes.guess_type(safe_name)[0] or suffix.lstrip(".")).strip(),
        remark=(payload.remark or "").strip(),
        source=SOURCE_AUDIO_MAKEUP,
        auto_checkin_by_whitelist=False,
        uploaded_on=now,
        uploaded_date=target_date,
        is_deleted=False,
    )
    db.add(row)
    try:
        await db.flush()
    except IntegrityError as exc:
        await db.rollback()
        raise HTTPException(status_code=400, detail="该读书内容已完成打卡，请勿重复提交。") from exc
    await db.refresh(row)
    # 补卡也算"当日打卡"——dedupe 走 uploaded_date 保证每日仅一次入账
    try:
        checkin_date = row.uploaded_date if isinstance(row.uploaded_date, date) else target_date
        await grant_points(
            db,
            user_id=user.id,
            rule_code="reading_checkin",
            business_type="audio_upload",
            business_id=int(row.id),
            dedupe_extra=f"d{checkin_date.isoformat()}",
            remark=f"读书补卡 {checkin_date.isoformat()}",
        )
        await record_reading_streak(db, user_id=user.id, checkin_date=checkin_date)
    except Exception:  # noqa: BLE001
        pass
    return _serialize_audio_record(row)


@router.delete("/my/audios/{audio_id}")
async def delete_my_audio(
    audio_id: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, bool]:
    row = await db.get(MagicAudioUpload, audio_id)
    if not row or row.user_id != user.id:
        raise HTTPException(status_code=404, detail="录音不存在。")
    row.is_deleted = True
    row.deleted_at = _now()
    await db.flush()
    return {"success": True}


@router.get("/my/audios/calendar")
async def get_my_audio_calendar(
    month: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> dict[str, Any]:
    whitelist_permissions = await get_user_whitelist_permissions(db, user.id)
    await _ensure_auto_audio_checkin(db, user, whitelist_permissions)
    month_start, _ = _parse_month(month)
    month_last_day_value = _month_last_day(month_start)
    result = await db.execute(
        select(MagicAudioUpload)
        .where(
            MagicAudioUpload.user_id == user.id,
            MagicAudioUpload.is_deleted.is_(False),
            MagicAudioUpload.uploaded_date >= month_start,
            MagicAudioUpload.uploaded_date <= month_last_day_value,
        )
        .order_by(MagicAudioUpload.uploaded_on.asc())
    )
    uploads = result.scalars().all()
    return {
        "month": month_start.strftime("%Y-%m"),
        "days": _build_audio_calendar_payload(month_start, month_last_day_value, uploads),
    }


async def _build_audio_stats(
    db: AsyncSession,
    month_text: str | None,
    department: str | None,
    user_id: int | None,
) -> list[dict[str, Any]]:
    month_start, month_end = _parse_month(month_text)
    expected = _expected_days(month_start, month_end)
    user_stmt = select(User).where(User.role.in_(["user", "admin"]), User.disabled.is_(False))
    if department:
        user_stmt = user_stmt.where(User.department == department)
    if user_id:
        user_stmt = user_stmt.where(User.id == user_id)
    user_stmt = user_stmt.order_by(User.id.asc())
    user_result = await db.execute(user_stmt)
    users = user_result.scalars().all()
    if not users:
        return []
    user_ids = [item.id for item in users]
    upload_result = await db.execute(
        select(MagicAudioUpload)
        .where(
            MagicAudioUpload.user_id.in_(user_ids),
            MagicAudioUpload.is_deleted.is_(False),
            MagicAudioUpload.uploaded_date >= month_start,
            MagicAudioUpload.uploaded_date <= month_end,
        )
        .order_by(MagicAudioUpload.uploaded_on.asc())
    )
    uploads = upload_result.scalars().all()
    grouped: dict[int, list[MagicAudioUpload]] = {}
    for item in uploads:
        grouped.setdefault(item.user_id, []).append(item)
    rows = []
    for target in users:
        items = grouped.get(target.id, [])
        upload_days = {item.uploaded_date.isoformat() for item in items if item.uploaded_date}
        upload_count = len(items)
        makeup_count = sum(1 for item in items if (item.source or "") == SOURCE_AUDIO_MAKEUP)
        missing = max(expected - len(upload_days), 0)
        rows.append({
            "user_id": target.id,
            "name": _user_name(target),
            "department": target.department or "",
            "month": month_start.strftime("%Y-%m"),
            "expected_upload_days": expected,
            "actual_upload_days": len(upload_days),
            "actual_upload_count": upload_count,
            "makeup_count": makeup_count,
            "missing_count": missing,
            "upload_rate": round((len(upload_days) / expected) * 100, 2) if expected > 0 else 0,
            "last_upload_time": _iso(items[-1].uploaded_on) if items else None,
        })
    return rows


@router.get("/admin/audio-stats")
async def get_audio_stats(
    month: str | None = None,
    department: str | None = None,
    user_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> list[dict[str, Any]]:
    del admin
    return await _build_audio_stats(db, month, department, user_id)


@router.get("/admin/audios/calendar")
async def get_admin_audio_calendar(
    month: str | None = None,
    department: str | None = None,
    user_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    month_start, _ = _parse_month(month)
    month_last_day_value = _month_last_day(month_start)
    reveal_whitelist = is_super_admin(admin)
    user_stmt = select(User).where(User.role.in_(["user", "admin"]), User.disabled.is_(False))
    if department:
        user_stmt = user_stmt.where(User.department == department)
    if user_id:
        user_stmt = user_stmt.where(User.id == user_id)
    user_stmt = user_stmt.order_by(User.id.asc())
    user_result = await db.execute(user_stmt)
    users = user_result.scalars().all()
    user_map = {item.id: item for item in users}
    user_ids = list(user_map)
    uploads: list[MagicAudioUpload] = []
    if user_ids:
        upload_result = await db.execute(
            select(MagicAudioUpload)
            .where(
                MagicAudioUpload.user_id.in_(user_ids),
                MagicAudioUpload.is_deleted.is_(False),
                MagicAudioUpload.uploaded_date >= month_start,
                MagicAudioUpload.uploaded_date <= month_last_day_value,
            )
            .order_by(MagicAudioUpload.uploaded_on.asc())
        )
        uploads = upload_result.scalars().all()
    return {
        "month": month_start.strftime("%Y-%m"),
        "user_id": user_id,
        "department": department or "",
        "scope": "user" if user_id else "all",
        "days": _build_audio_calendar_payload(
            month_start,
            month_last_day_value,
            uploads,
            user_map,
            aggregate_users=not user_id,
            reveal_whitelist=reveal_whitelist,
        ),
    }


@router.get("/admin/audio-stats/export")
async def export_audio_stats(
    month: str | None = None,
    department: str | None = None,
    user_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> StreamingResponse:
    del admin
    rows = await _build_audio_stats(db, month, department, user_id)
    export_rows = [
        [
            item["name"],
            item["department"],
            item["month"],
            item["expected_upload_days"],
            item["actual_upload_days"],
            item["actual_upload_count"],
            item["missing_count"],
            item["upload_rate"],
            item["last_upload_time"] or "",
        ]
        for item in rows
    ]
    return _xlsx_response(
        f"magic_audio_stats_{month or date.today().strftime('%Y-%m')}.xlsx",
        ["姓名", "部门", "月份", "应上传天数", "实际上传天数", "实际上传次数", "缺少次数", "上传率", "最后上传时间"],
        export_rows,
    )


@router.get("/admin/audio-statistics/reading-contents")
async def get_admin_reading_content_statistics(
    month: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    reading_content_id: int | None = None,
    department: str | None = None,
    user_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    contents = await _list_filtered_reading_contents_for_admin(
        db,
        admin=admin,
        month=month,
        start_date=start_date,
        end_date=end_date,
        reading_content_id=reading_content_id,
    )
    users = await _list_filtered_employee_users(db, department=department, user_id=user_id)
    expected_map = await _build_content_expected_users_map(db, contents=contents, users=users)
    uploads_map = await _build_uploads_map_for_contents(
        db,
        content_ids=[int(item.id) for item in contents],
        user_ids=[int(item.id) for item in users],
    )
    targets_map = await _build_admin_reading_targets_map(db, [int(item.id) for item in contents])
    now = _now()
    rows: list[dict[str, Any]] = []
    for content in contents:
        targets = targets_map.get(int(content.id), [])
        expected_users = expected_map.get(int(content.id), [])
        expected_count = len(expected_users)
        completed_count = 0
        status_codes: set[str] = set()
        for target in expected_users:
            upload = uploads_map.get((int(content.id), int(target.id)))
            if upload is not None:
                completed_count += 1
            status_code, _status_text = _status_code_for_user_content(content, upload, now=now)
            status_codes.add(status_code)
        if status and status != "all":
            if status not in status_codes:
                continue
        pending_count = max(expected_count - completed_count, 0)
        rows.append({
            "reading_content_id": int(content.id),
            "reading_date": _iso(content.reading_date),
            "push_at": _iso(_effective_push_at(content)),
            "push_time": content.push_time.isoformat() if content.push_time else _effective_push_at(content).time().isoformat(),
            "title": content.title,
            "target_summary": _reading_target_summary(targets),
            "expected_count": expected_count,
            "completed_count": completed_count,
            "pending_count": pending_count,
            "completion_rate": round((completed_count / expected_count) * 100, 2) if expected_count else 0,
            "makeup_deadline_at": _iso(content.makeup_deadline_at),
            "is_deadline_passed": bool(content.makeup_deadline_at and now > content.makeup_deadline_at),
            "has_checkins": completed_count > 0,
        })
    legacy_count_result = await db.execute(
        select(func.count(MagicAudioUpload.id)).where(
            MagicAudioUpload.reading_content_id.is_(None),
            MagicAudioUpload.is_deleted.is_(False),
        )
    )
    legacy_count = int(legacy_count_result.scalar_one() or 0)
    return {
        "items": rows,
        "legacy_unbound_hint": f"存在 {legacy_count} 条历史未绑定录音，未计入具体读书内容完成率。" if legacy_count > 0 else "",
    }


@router.get("/admin/audio-statistics/reading-contents/{reading_content_id}/users")
async def get_admin_reading_content_user_statistics(
    reading_content_id: int,
    department: str | None = None,
    user_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    content, rows, legacy_hint = await _build_reading_content_user_rows(
        db,
        admin=admin,
        reading_content_id=reading_content_id,
        department=department,
        user_id=user_id,
        status=status,
    )
    return {
        "reading_content_id": int(content.id),
        "title": content.title,
        "reading_date": _iso(content.reading_date),
        "push_at": _iso(_effective_push_at(content)),
        "makeup_deadline_at": _iso(content.makeup_deadline_at),
        "items": rows,
        "legacy_unbound_hint": legacy_hint,
    }


@router.get("/admin/audio-statistics/reading-contents/export")
async def export_admin_reading_content_statistics(
    month: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    reading_content_id: int | None = None,
    department: str | None = None,
    user_id: int | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> StreamingResponse:
    row_contexts = await _collect_admin_reading_audio_export_rows(
        db,
        admin=admin,
        month=month,
        start_date=start_date,
        end_date=end_date,
        reading_content_id=reading_content_id,
        department=department,
        user_id=user_id,
        status=status,
    )
    return _build_reading_audio_export_workbook(
        filename=_build_reading_audio_export_filename(month),
        columns=_resolve_reading_audio_export_columns(None),
        row_contexts=row_contexts,
    )


@router.post("/admin/audio-statistics/reading-contents/export")
async def export_admin_reading_content_statistics_with_columns(
    payload: AdminReadingAudioStatisticsExportPayload = Body(default_factory=AdminReadingAudioStatisticsExportPayload),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> StreamingResponse:
    row_contexts = await _collect_admin_reading_audio_export_rows(
        db,
        admin=admin,
        month=payload.month,
        start_date=payload.start_date,
        end_date=payload.end_date,
        reading_content_id=payload.reading_content_id,
        department=payload.department,
        user_id=payload.user_id,
        status=payload.status,
    )
    return _build_reading_audio_export_workbook(
        filename=_build_reading_audio_export_filename(payload.month),
        columns=_resolve_reading_audio_export_columns(payload.columns),
        row_contexts=row_contexts,
    )
