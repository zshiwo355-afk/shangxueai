from __future__ import annotations

import asyncio
import base64
import json
import logging
import mimetypes
import posixpath
import re
from collections.abc import Iterable
from datetime import date, datetime, time, timedelta
from io import BytesIO
from typing import Any
from uuid import uuid4
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from fastapi import Body, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import RedirectResponse
from sqlalchemy import and_, case, delete as sql_delete, desc, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..access import is_super_admin
from ..auth import get_current_user, require_admin
from ..db import get_db, session_scope
from ..magic_academy_schemas import (
    ReadingContentImportConfirmPayload,
    ReadingContentImportJobResponse,
    ReadingContentImportStartResponse,
    ReadingSeriesPayload,
)
from ..magic_auto_actions import enqueue_audio_actions_for_reading_content
from ..magic_push_service import (
    batch_to_dict,
    entries_to_dicts,
    get_latest_batch,
    get_push_entries,
    run_reading_manual_retry,
)
from ..models import ConfigOption, MagicAudioUpload, MagicReadingContent, MagicReadingContentTarget, MagicReadingSeries, MagicReadingSeriesTarget, MaterialAsset, User, UserWhitelist
from . import router
from ._oss import (
    _build_oss_bucket,  # CODEX_MODIFIED
    _build_oss_object_url,
    _build_signed_stream_url,
    _ensure_oss_settings,
    _upload_binary_to_oss,
    _validate_reading_image_payload,
)
from ._resource_cleanup import schedule_oss_object_cleanup
from ._utils import (
    READING_CONTENT_ACTIVE,
    SOURCE_AUDIO_MAKEUP,
    SOURCE_AUDIO_USER_UPLOAD,
    SOURCE_WHITELIST_AUTO,
    _iso,
    _json_loads,
    _normalize_image_source,
    _normalize_reading_target_type,
    _now,
    _parse_form_id_list,
    _parse_month,
    _safe_filename,
    _user_department,
    _user_name,
    _user_position,
    _xlsx_response,
)
from ._video_helpers import _get_material_asset_or_403

logger = logging.getLogger("app.magic_academy_api.reading")

READING_STATUS_DISABLED = "disabled"
READING_SERIES_STATUSES = {"draft", "active", "paused", "archived"}
READING_SERIES_ARCHIVED = "archived"
READING_IMPORT_TARGET_TYPE_ALIASES = {
    "全员": "all",
    "all": "all",
    "全部员工": "all",
    "仅新人": "all_newcomers",
    "新人": "all_newcomers",
    "all_newcomers": "all_newcomers",
    "部门": "department",
    "department": "department",
    "员工": "user",
    "指定员工": "user",
    "user": "user",
    "岗位": "position",
    "position": "position",
    "职级": "job_level",
    "job_level": "job_level",
    "M线": "job_level",
    "P线": "job_level",
    "在职状态": "employment_status",
    "employment_status": "employment_status",
}
READING_IMPORT_PREVIEW_TTL_SECONDS = 1800
READING_IMPORT_JOB_TTL_SECONDS = 3600
_reading_import_preview_cache: dict[str, dict[str, Any]] = {}
_reading_import_job_cache: dict[str, dict[str, Any]] = {}
_reading_import_job_tasks: dict[str, asyncio.Task] = {}
_reading_import_state_lock = asyncio.Lock()


def _reading_target_to_dict(item: MagicReadingContentTarget) -> dict[str, Any]:
    return {
        "id": int(item.id),
        "target_type": item.target_type,
        "target_id": item.target_id or "",
    }


def _reading_series_target_to_dict(item: MagicReadingSeriesTarget) -> dict[str, Any]:
    return {
        "id": int(item.id),
        "target_type": item.target_type,
        "target_id": item.target_id or "",
    }


def _is_content_out_of_series_range(content: MagicReadingContent, series: MagicReadingSeries) -> bool:
    if series.start_date and content.reading_date < series.start_date:
        return True
    if series.end_date and content.reading_date > series.end_date:
        return True
    return False


def _reading_series_to_dict(
    item: MagicReadingSeries,
    *,
    content_count: int = 0,
    enabled_content_count: int = 0,
    out_of_range_content_count: int = 0,
    targets: list[MagicReadingSeriesTarget] | None = None,
    target_summary: str = "未设置",
    contents: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    data = {
        "id": int(item.id),
        "title": item.title,
        "description": item.description or "",
        "start_date": _iso(item.start_date),
        "end_date": _iso(item.end_date),
        "status": item.status or "draft",
        "content_count": int(content_count or 0),
        "enabled_content_count": int(enabled_content_count or 0),
        "out_of_range_content_count": int(out_of_range_content_count or 0),
        "target_summary": target_summary,
        "targets": [_reading_series_target_to_dict(target) for target in (targets or [])],
        "created_by": int(item.created_by),
        "created_at": _iso(item.created_at),
        "updated_at": _iso(item.updated_at),
    }
    if contents is not None:
        data["contents"] = contents
    return data


async def _get_reading_series_or_404(db: AsyncSession, series_id: int) -> MagicReadingSeries:
    row = await db.get(MagicReadingSeries, series_id)
    if not row:
        raise HTTPException(status_code=404, detail="读书系列不存在。")
    return row


def _can_manage_reading_series(admin: User, row: MagicReadingSeries) -> bool:
    if is_super_admin(admin):
        return True
    if not row.created_by:
        return True
    return int(row.created_by) == int(admin.id)


def _reading_series_targets_summary(targets: list[MagicReadingSeriesTarget]) -> str:
    if not targets:
        return "未设置"
    if any((item.target_type or "").lower() == "all" for item in targets):
        return "全部员工"
    departments = [item.target_id for item in targets if item.target_type == "department" and item.target_id]
    positions = [item.target_id for item in targets if item.target_type == "position" and item.target_id]
    job_levels = [item.target_id for item in targets if item.target_type == "job_level" and item.target_id]
    employment_statuses = [item.target_id for item in targets if item.target_type == "employment_status" and item.target_id]
    users = [item.target_id for item in targets if item.target_type == "user" and item.target_id]
    parts: list[str] = []
    if departments:
        parts.append("部门：" + "、".join(departments[:2]) + (f"等 {len(departments)} 个" if len(departments) > 2 else ""))
    if positions:
        parts.append("岗位：" + "、".join(positions[:2]) + (f"等 {len(positions)} 个" if len(positions) > 2 else ""))
    if job_levels:
        parts.append("职级：" + "、".join(job_levels[:2]) + (f"等 {len(job_levels)} 个" if len(job_levels) > 2 else ""))
    if employment_statuses:
        parts.append("在职状态：" + "、".join(employment_statuses[:2]) + (f"等 {len(employment_statuses)} 个" if len(employment_statuses) > 2 else ""))
    if users:
        parts.append(f"人员 {len(users)} 人")
    return "；".join(parts) or "未设置"


async def _get_reading_series_targets_map(
    db: AsyncSession,
    series_ids: list[int],
) -> dict[int, list[MagicReadingSeriesTarget]]:
    if not series_ids:
        return {}
    rows = (
        await db.execute(
            select(MagicReadingSeriesTarget)
            .where(MagicReadingSeriesTarget.series_id.in_(series_ids))
            .order_by(MagicReadingSeriesTarget.id.asc())
        )
    ).scalars().all()
    result: dict[int, list[MagicReadingSeriesTarget]] = {}
    for row in rows:
        result.setdefault(int(row.series_id), []).append(row)
    return result


async def _replace_reading_series_targets(
    db: AsyncSession,
    series_id: int,
    targets: list[Any],
) -> None:
    await db.execute(sql_delete(MagicReadingSeriesTarget).where(MagicReadingSeriesTarget.series_id == series_id))
    rows: list[MagicReadingSeriesTarget] = []
    seen: set[tuple[str, str]] = set()
    for target in targets or []:
        target_type = (getattr(target, "target_type", "") or "").strip().lower()
        target_id = str(getattr(target, "target_id", "") or "").strip()
        if target_type == "all":
            target_id = ""
        if target_type not in {"all", "department", "position", "job_level", "employment_status", "user"}:
            continue
        if target_type != "all" and not target_id:
            continue
        key = (target_type, target_id)
        if key in seen:
            continue
        seen.add(key)
        rows.append(MagicReadingSeriesTarget(series_id=series_id, target_type=target_type, target_id=target_id))
    if rows:
        db.add_all(rows)


async def _validate_reading_series_id(
    db: AsyncSession,
    admin: User,
    series_id: int | None,
) -> int | None:
    if not series_id:
        return None
    row = await _get_reading_series_or_404(db, int(series_id))
    if not _can_manage_reading_series(admin, row):
        raise HTTPException(status_code=403, detail="无权使用该读书系列。")
    if (row.status or "draft") == READING_SERIES_ARCHIVED:
        raise HTTPException(status_code=400, detail="已归档系列不能继续关联新内容。")
    return int(row.id)


async def _assert_reading_date_allowed_by_series(
    db: AsyncSession,
    series_id: int | None,
    reading_date: date,
) -> None:
    if not series_id:
        return
    series = await db.get(MagicReadingSeries, int(series_id))
    if not series:
        return
    if series.start_date and reading_date < series.start_date:
        raise HTTPException(status_code=400, detail="读书内容日期不能早于所属系列计划周期。")
    if series.end_date and reading_date > series.end_date:
        raise HTTPException(status_code=400, detail="读书内容日期不能晚于所属系列计划周期。")


def _reading_image_url(object_key: str) -> str:
    if not (object_key or "").strip():
        return ""
    return _build_signed_stream_url(object_key.strip())


def _safe_reading_image_url(object_key: str, fallback: str = "") -> str:
    if not (object_key or "").strip():
        return fallback or ""
    try:
        return _reading_image_url(object_key)
    except Exception:
        return fallback or ""


def _is_empty_import_row(values: list[Any], embedded_image: dict[str, Any] | None = None) -> bool:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return False
            continue
        return False
    return embedded_image is None


def _default_push_at(reading_date: date) -> datetime:
    return datetime.combine(reading_date, time(hour=18, minute=30, second=0))


def _effective_push_at(item: MagicReadingContent) -> datetime:
    return item.push_at or _default_push_at(item.reading_date)


def _default_makeup_deadline(push_at: datetime) -> datetime:
    deadline_date = push_at.date() + timedelta(days=2)
    return datetime.combine(deadline_date, time(hour=23, minute=59, second=59))


def _resolve_makeup_deadline(
    reading_date: date,
    push_at: datetime | None,
    makeup_deadline_at: datetime | None,
) -> datetime | None:
    if makeup_deadline_at is not None:
        return makeup_deadline_at
    if push_at is None:
        push_at = _default_push_at(reading_date)
    return _default_makeup_deadline(push_at)


def _parse_push_time_text(value: Any) -> time:
    text = str(value or "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="推送时间不能为空。")
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail="推送时间格式应为 HH:MM 或 HH:MM:SS。")


def _parse_datetime_text(value: Any, *, field_name: str) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail=f"{field_name}格式应为 YYYY-MM-DD HH:MM。")


def _cleanup_reading_import_state() -> None:
    now = _now()
    preview_expired = [
        key for key, item in _reading_import_preview_cache.items()
        if (now - item["created_at"]).total_seconds() > READING_IMPORT_PREVIEW_TTL_SECONDS
    ]
    for key in preview_expired:
        _reading_import_preview_cache.pop(key, None)
    job_expired = [
        key for key, item in _reading_import_job_cache.items()
        if (now - item["created_at"]).total_seconds() > READING_IMPORT_JOB_TTL_SECONDS
    ]
    for key in job_expired:
        _reading_import_job_cache.pop(key, None)
        task = _reading_import_job_tasks.pop(key, None)
        if task and task.done():
            _ = task.exception() if not task.cancelled() else None


def _serialize_preview_row_for_response(row: dict[str, Any]) -> dict[str, Any]:
    parsed = dict(row.get("parsed") or {})
    parsed.pop("embedded_image_base64", None)
    return {
        "row_number": row.get("row_number"),
        "raw": row.get("raw") or {},
        "parsed": parsed,
        "errors": list(row.get("errors") or []),
        "warnings": list(row.get("warnings") or []),
        "can_import": bool(row.get("can_import")),
    }


async def _get_auto_checkin_whitelist_user_ids(db: AsyncSession) -> set[int]:
    result = await db.execute(
        select(UserWhitelist.user_id)
        .join(User, User.id == UserWhitelist.user_id)
        .where(
            User.role.in_(["user", "admin"]),
            User.disabled.is_(False),
            UserWhitelist.enabled.is_(True),
            UserWhitelist.auto_checkin_enabled.is_(True),
        )
    )
    return {int(item[0]) for item in result.all()}


async def _update_reading_import_job(job_id: str, **updates: Any) -> None:
    async with _reading_import_state_lock:
        job = _reading_import_job_cache.get(job_id)
        if not job:
            return
        job.update(updates)
        job["updated_at"] = _now()


async def _run_reading_import_job(job_id: str, rows: list[dict[str, Any]], admin_id: int) -> None:
    await _update_reading_import_job(job_id, status="running")
    try:
        async with session_scope() as db:
            admin = await db.get(User, admin_id)
            if not admin:
                raise RuntimeError("创建导入任务的管理员不存在。")
            whitelist_user_ids = await _get_auto_checkin_whitelist_user_ids(db)
            for index, raw_row in enumerate(rows, start=1):
                try:
                    import_image_source = str(raw_row.get("image_source") or "upload")
                    if (
                        import_image_source == "url"
                        and not str(raw_row.get("image_url") or "").strip()
                        and not raw_row.get("embedded_image_base64")
                        and not raw_row.get("material_asset_id")
                    ):
                        import_image_source = "upload"
                    await _create_reading_content_record(
                        db,
                        admin=admin,
                        reading_date=_parse_date_value(raw_row.get("reading_date")),
                        push_time_value=_parse_push_time_text(raw_row.get("push_time")),
                        title=str(raw_row.get("title") or ""),
                        description=str(raw_row.get("description") or ""),
                        image_source=import_image_source,
                        material_asset_id=int(raw_row["material_asset_id"]) if raw_row.get("material_asset_id") else None,
                        series_id=int(raw_row["series_id"]) if raw_row.get("series_id") else None,
                        target_type=str(raw_row.get("target_type") or "user"),
                        target_user_ids=[int(v) for v in (raw_row.get("target_user_ids") or [])],
                        target_department_ids=[str(v).strip() for v in (raw_row.get("target_department_ids") or []) if str(v).strip()],
                        target_position_ids=[str(v).strip() for v in (raw_row.get("target_position_ids") or []) if str(v).strip()],
                        target_job_level_ids=[str(v).strip() for v in (raw_row.get("target_job_level_ids") or []) if str(v).strip()],
                        target_employment_status_ids=[str(v).strip() for v in (raw_row.get("target_employment_status_ids") or []) if str(v).strip()],
                        makeup_deadline_at=_parse_datetime_text(raw_row.get("makeup_deadline_at"), field_name="补卡截止时间"),
                        image=None,
                        image_url_text=str(raw_row.get("image_url") or ""),
                        embedded_image_bytes=base64.b64decode(raw_row["embedded_image_base64"]) if raw_row.get("embedded_image_base64") else None,
                        embedded_image_name=str(raw_row.get("embedded_image_name") or ""),
                        embedded_image_mime_type=str(raw_row.get("embedded_image_mime_type") or ""),
                        auto_checkin_whitelist_user_ids=whitelist_user_ids,
                    )
                    await db.commit()
                    async with _reading_import_state_lock:
                        job = _reading_import_job_cache.get(job_id)
                        if not job:
                            return
                        job["processed"] = index
                        job["success_count"] += 1
                        job["updated_at"] = _now()
                except Exception as exc:  # noqa: BLE001
                    await db.rollback()
                    async with _reading_import_state_lock:
                        job = _reading_import_job_cache.get(job_id)
                        if not job:
                            return
                        job["processed"] = index
                        job["failure_count"] += 1
                        job["failures"].append({
                            "row_number": raw_row.get("row_number"),
                            "title": str(raw_row.get("title") or ""),
                            "reason": str(getattr(exc, "detail", "") or exc),
                        })
                        job["updated_at"] = _now()
        final_status = "completed"
        if _reading_import_job_cache.get(job_id, {}).get("failure_count"):
            final_status = "completed_with_errors"
        await _update_reading_import_job(job_id, status=final_status)
    except Exception as exc:  # noqa: BLE001
        await _update_reading_import_job(job_id, status="failed", error=str(exc))
    finally:
        async with _reading_import_state_lock:
            _reading_import_job_tasks.pop(job_id, None)


def _split_multi_text(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, list):
        return [str(item or "").strip() for item in value if str(item or "").strip()]
    text = str(value or "").strip()
    if not text:
        return []
    return [item.strip() for item in text.replace("，", ",").split(",") if item.strip()]


def _parse_date_value(value: Any, *, field_name: str = "日期") -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value or "").strip()
    if not text:
        raise ValueError(f"{field_name}不能为空。")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field_name}格式应为 YYYY-MM-DD。")


def _parse_time_value(value: Any) -> time:
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time().replace(microsecond=0)
    text = str(value or "").strip()
    if not text:
        raise ValueError("推送时间不能为空。")
    for fmt in ("%H:%M", "%H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    raise ValueError("推送时间格式应为 HH:MM。")


def _parse_excel_datetime(value: Any, *, field_name: str) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    text = str(value or "").strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    raise ValueError(f"{field_name}格式应为 YYYY-MM-DD HH:MM。")


def _normalize_import_target_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    normalized = READING_IMPORT_TARGET_TYPE_ALIASES.get(text)
    if normalized:
        return normalized
    raise ValueError("目标人群类型不合法。")


def _build_push_at(reading_date: date, push_time: time) -> datetime:
    return datetime.combine(reading_date, push_time)


def _reading_content_to_dict(
    item: MagicReadingContent,
    *,
    targets: list[MagicReadingContentTarget] | None = None,
    image_url: str | None = None,
    creator: User | None = None,
    series: MagicReadingSeries | None = None,
    push_count: int | None = None,
    is_locked: bool = False,
    current_status: str | None = None,
    completed: bool = False,
    upload_id: int | None = None,
    can_submit: bool | None = None,
    submit_disabled_reason: str = "",
    completed_count: int | None = None,
) -> dict[str, Any]:
    resolved_url = image_url if image_url is not None else (item.image_url or "")
    target_rows = list(targets or [])
    push_at = _effective_push_at(item)
    makeup_deadline_at = item.makeup_deadline_at
    return {
        "id": int(item.id),
        "series_id": int(item.series_id) if item.series_id else None,
        "series_title": series.title if series else "",
        "reading_date": _iso(item.reading_date),
        "push_time": item.push_time.isoformat() if item.push_time else push_at.time().isoformat(),
        "push_at": _iso(push_at),
        "makeup_deadline_at": _iso(makeup_deadline_at),
        "title": item.title,
        "description": item.description or "",
        "source_type": (item.source_type or "upload").strip().lower(),
        "material_asset_id": int(item.material_asset_id) if item.material_asset_id else None,
        "image_object_key": item.image_object_key or "",
        "image_url": resolved_url,
        "image_file_name": item.image_file_name or "",
        "image_mime_type": item.image_mime_type or "",
        "image_size": int(item.image_size or 0),
        "status": item.status or READING_CONTENT_ACTIVE,
        "created_by": int(item.created_by),
        "creator_name": _user_name(creator) if creator else "",
        "created_at": _iso(item.created_at),
        "updated_at": _iso(item.updated_at),
        "targets": [_reading_target_to_dict(target) for target in target_rows],
        "push_count": int(push_count if push_count is not None else sum(1 for target in target_rows if target.target_type == "user")),
        "completed_count": int(completed_count or 0),
        "pending_count": max(int(push_count if push_count is not None else 0) - int(completed_count or 0), 0),
        "completion_rate": round((int(completed_count or 0) / int(push_count)) * 100, 2) if push_count else 0,
        "has_checkins": bool(is_locked),
        "is_locked": bool(is_locked),
        "editable": True,
        "core_fields_locked": bool(is_locked),
        "deletable": not is_locked,
        "edit_lock_reason": "该读书内容已有打卡记录，不能修改日期、标题、图片、派发对象等核心字段，请停用后重新创建。" if is_locked else "",
        "delete_disabled_reason": "该内容已有打卡记录，不允许删除，请使用停用。" if is_locked else "",
        "current_status": current_status or "",
        "completed": bool(completed),
        "upload_id": int(upload_id) if upload_id else None,
        "record_id": int(upload_id) if upload_id else None,
        "reading_content_id": int(item.id),
        "can_submit": can_submit,
        "submit_disabled_reason": submit_disabled_reason,
    }


async def _get_reading_content_targets_map(
    db: AsyncSession,
    content_ids: list[int],
) -> dict[int, list[MagicReadingContentTarget]]:
    if not content_ids:
        return {}
    result = await db.execute(
        select(MagicReadingContentTarget)
        .where(MagicReadingContentTarget.content_id.in_(content_ids))
        .order_by(MagicReadingContentTarget.id.asc())
    )
    mapping: dict[int, list[MagicReadingContentTarget]] = {}
    for item in result.scalars().all():
        mapping.setdefault(item.content_id, []).append(item)
    return mapping


def _reading_target_matches_user(user: User, target: MagicReadingContentTarget) -> bool:
    ttype = (target.target_type or "").strip().lower()
    target_id = (target.target_id or "").strip()
    is_employee_role = user.role in {"user", "admin"}
    if ttype == "all":
        return is_employee_role
    if ttype == "all_newcomers":
        return is_employee_role and bool(user.is_newcomer)
    if ttype == "department":
        return is_employee_role and _user_department(user) == target_id
    if ttype == "position":
        return is_employee_role and _user_position(user) == target_id
    if ttype == "job_level":
        return is_employee_role and (user.job_level or "M线").strip() == target_id
    if ttype == "employment_status":
        return is_employee_role and (user.employment_status or "").strip() == target_id
    if ttype == "user":
        return str(user.id) == target_id
    return False


async def _get_reading_content_or_404(
    db: AsyncSession,
    content_id: int,
) -> MagicReadingContent:
    row = await db.get(MagicReadingContent, content_id)
    if not row or row.is_deleted:
        raise HTTPException(status_code=404, detail="读书内容不存在。")
    return row


def _can_manage_reading_content(admin: User, row: MagicReadingContent) -> bool:
    if is_super_admin(admin):
        return True
    if not row.created_by:
        return True
    return int(row.created_by) == int(admin.id)


async def _replace_reading_targets(
    db: AsyncSession,
    content_id: int,
    *,
    target_type: str,
    user_ids: list[int],
    department_names: list[str],
    position_names: list[str],
    job_level_names: list[str] | None = None,
    employment_status_values: list[str] | None = None,
) -> list[MagicReadingContentTarget]:
    job_level_names = job_level_names or []
    employment_status_values = employment_status_values or []
    await db.execute(sql_delete(MagicReadingContentTarget).where(MagicReadingContentTarget.content_id == content_id))
    rows: list[MagicReadingContentTarget] = []
    if target_type == "all":
        rows.append(MagicReadingContentTarget(content_id=content_id, target_type="all", target_id="0"))
    elif target_type == "all_newcomers":
        rows.append(MagicReadingContentTarget(content_id=content_id, target_type="all_newcomers", target_id="1"))
    elif target_type == "department":
        rows.extend(
            MagicReadingContentTarget(content_id=content_id, target_type="department", target_id=name)
            for name in department_names
        )
    elif target_type == "position":
        rows.extend(
            MagicReadingContentTarget(content_id=content_id, target_type="position", target_id=name)
            for name in position_names
        )
    elif target_type == "job_level":
        rows.extend(
            MagicReadingContentTarget(content_id=content_id, target_type="job_level", target_id=name)
            for name in job_level_names
        )
    elif target_type == "employment_status":
        rows.extend(
            MagicReadingContentTarget(content_id=content_id, target_type="employment_status", target_id=value)
            for value in employment_status_values
        )
    else:
        rows.extend(
            MagicReadingContentTarget(content_id=content_id, target_type="user", target_id=str(user_id))
            for user_id in user_ids
        )
    for row in rows:
        db.add(row)
    await db.flush()
    return rows


def _normalize_explicit_reading_targets(raw_targets: Any) -> list[dict[str, str]]:
    if not isinstance(raw_targets, list):
        return []
    result: list[dict[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for target in raw_targets:
        if not isinstance(target, dict):
            continue
        target_type = str(target.get("target_type") or "").strip().lower()
        target_id = str(target.get("target_id") or "").strip()
        if target_type == "all":
            target_id = "0"
        if target_type == "all_newcomers":
            target_id = "1"
        if target_type not in {"all", "all_newcomers", "department", "position", "job_level", "employment_status", "user"}:
            continue
        if target_type != "all" and not target_id:
            continue
        key = (target_type, target_id)
        if key in seen:
            continue
        seen.add(key)
        result.append({"target_type": target_type, "target_id": target_id})
    return result


async def _replace_reading_targets_from_payload(
    db: AsyncSession,
    content_id: int,
    raw_targets: Any,
) -> list[MagicReadingContentTarget]:
    targets = _normalize_explicit_reading_targets(raw_targets)
    if not targets:
        raise HTTPException(status_code=400, detail="请选择至少一个派发对象。")
    await db.execute(sql_delete(MagicReadingContentTarget).where(MagicReadingContentTarget.content_id == content_id))
    rows = [
        MagicReadingContentTarget(content_id=content_id, target_type=item["target_type"], target_id=item["target_id"])
        for item in targets
    ]
    db.add_all(rows)
    await db.flush()
    return rows


async def _validate_reading_recipients(
    db: AsyncSession,
    *,
    target_type: str,
    target_user_ids: list[int],
    target_department_names: list[str],
    target_position_names: list[str],
    target_job_level_names: list[str] | None = None,
    target_employment_status_values: list[str] | None = None,
) -> tuple[list[int], list[str], list[str], int]:
    target_job_level_names = target_job_level_names or []
    target_employment_status_values = target_employment_status_values or []
    target_type = _normalize_reading_target_type(target_type)
    if target_type == "all":
        result = await db.execute(select(func.count(User.id)).where(User.role.in_(["user", "admin"]), User.disabled.is_(False)))
        return [], [], [], int(result.scalar_one() or 0)
    if target_type == "all_newcomers":
        result = await db.execute(
            select(func.count(User.id)).where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.is_newcomer.is_(True))
        )
        return [], [], [], int(result.scalar_one() or 0)
    if target_type == "department":
        names = sorted({(name or "").strip() for name in target_department_names if (name or "").strip()})
        if not names:
            raise HTTPException(status_code=400, detail="请选择至少一个部门。")
        result = await db.execute(
            select(User.id, User.department)
            .where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.department.in_(names))
        )
        rows = result.all()
        matched_departments = sorted({(department or "").strip() for _, department in rows if (department or "").strip()})
        if not rows:
            raise HTTPException(status_code=400, detail="所选部门下没有可推送员工。")
        return [], matched_departments, [], len(rows)
    if target_type == "position":
        names = sorted({(name or "").strip() for name in target_position_names if (name or "").strip()})
        if not names:
            raise HTTPException(status_code=400, detail="请选择至少一个岗位。")
        result = await db.execute(
            select(User.id, User.position)
            .where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.position.in_(names))
        )
        rows = result.all()
        matched_positions = sorted({(position or "").strip() for _, position in rows if (position or "").strip()})
        if not rows:
            raise HTTPException(status_code=400, detail="所选岗位下没有可推送员工。")
        return [], [], matched_positions, len(rows)
    if target_type == "employment_status":
        values = sorted({(v or "").strip() for v in target_employment_status_values if (v or "").strip()})
        if not values:
            raise HTTPException(status_code=400, detail="请选择至少一个在职状态。")
        result = await db.execute(
            select(User.id, User.employment_status)
            .where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.employment_status.in_(values))
        )
        rows = result.all()
        if not rows:
            raise HTTPException(status_code=400, detail="所选在职状态下没有可推送员工。")
        return [], [], [], len(rows)
    if target_type == "job_level":
        names = sorted({(name or "").strip() for name in target_job_level_names if (name or "").strip()})
        if not names:
            raise HTTPException(status_code=400, detail="请选择至少一个职级。")
        invalid = [name for name in names if name not in {"M线", "P线"}]
        if invalid:
            raise HTTPException(status_code=400, detail="职级仅支持 M线 / P线。")
        result = await db.execute(
            select(User.id, User.job_level)
            .where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.job_level.in_(names))
        )
        rows = result.all()
        if not rows:
            raise HTTPException(status_code=400, detail="所选职级下没有可推送员工。")
        return [], [], [], len(rows)
    user_ids = sorted(set(target_user_ids))
    if not user_ids:
        raise HTTPException(status_code=400, detail="请选择至少一个员工。")
    result = await db.execute(select(User.id).where(User.id.in_(user_ids), User.role.in_(["user", "admin"]), User.disabled.is_(False)))
    existing_ids = sorted({int(item[0]) for item in result.all()})
    if len(existing_ids) != len(user_ids):
        raise HTTPException(status_code=400, detail="推送对象里包含无效员工。")
    return existing_ids, [], [], len(existing_ids)


async def _count_reading_targets(
    db: AsyncSession,
    targets: list[MagicReadingContentTarget],
) -> int:
    if not targets:
        return 0
    if any((item.target_type or "").lower() == "all" for item in targets):
        result = await db.execute(select(func.count(User.id)).where(User.role.in_(["user", "admin"]), User.disabled.is_(False)))
        return int(result.scalar_one() or 0)
    if any((item.target_type or "").lower() == "all_newcomers" for item in targets):
        result = await db.execute(
            select(func.count(User.id)).where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.is_newcomer.is_(True))
        )
        return int(result.scalar_one() or 0)
    filters = []
    departments = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "department" and (item.target_id or "").strip()
    })
    if departments:
        filters.append(User.department.in_(departments))
    positions = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "position" and (item.target_id or "").strip()
    })
    if positions:
        filters.append(User.position.in_(positions))
    job_levels = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "job_level" and (item.target_id or "").strip()
    })
    if job_levels:
        filters.append(User.job_level.in_(job_levels))
    employment_statuses = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "employment_status" and (item.target_id or "").strip()
    })
    if employment_statuses:
        filters.append(User.employment_status.in_(employment_statuses))
    user_ids = sorted({
        int(item.target_id)
        for item in targets
        if (item.target_type or "").lower() == "user" and str(item.target_id or "").isdigit()
    })
    if user_ids:
        filters.append(User.id.in_(user_ids))
    if not filters:
        return 0
    result = await db.execute(
        select(func.count(func.distinct(User.id))).where(User.role.in_(["user", "admin"]), User.disabled.is_(False), or_(*filters))
    )
    return int(result.scalar_one() or 0)


async def _collect_target_user_ids(
    db: AsyncSession,
    targets: list[MagicReadingContentTarget],
) -> list[int]:
    if not targets:
        return []
    if any((item.target_type or "").lower() == "all" for item in targets):
        result = await db.execute(select(User.id).where(User.role.in_(["user", "admin"]), User.disabled.is_(False)))
        return [int(item[0]) for item in result.all()]
    if any((item.target_type or "").lower() == "all_newcomers" for item in targets):
        result = await db.execute(
            select(User.id).where(User.role.in_(["user", "admin"]), User.disabled.is_(False), User.is_newcomer.is_(True))
        )
        return [int(item[0]) for item in result.all()]
    filters = []
    departments = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "department" and (item.target_id or "").strip()
    })
    if departments:
        filters.append(User.department.in_(departments))
    positions = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "position" and (item.target_id or "").strip()
    })
    if positions:
        filters.append(User.position.in_(positions))
    job_levels = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "job_level" and (item.target_id or "").strip()
    })
    if job_levels:
        filters.append(User.job_level.in_(job_levels))
    employment_statuses = sorted({
        (item.target_id or "").strip()
        for item in targets
        if (item.target_type or "").lower() == "employment_status" and (item.target_id or "").strip()
    })
    if employment_statuses:
        filters.append(User.employment_status.in_(employment_statuses))
    user_ids = sorted({
        int(item.target_id)
        for item in targets
        if (item.target_type or "").lower() == "user" and str(item.target_id or "").isdigit()
    })
    if user_ids:
        filters.append(User.id.in_(user_ids))
    if not filters:
        return []
    result = await db.execute(select(User.id).where(User.role.in_(["user", "admin"]), User.disabled.is_(False), or_(*filters)))
    return sorted({int(item[0]) for item in result.all()})


async def _count_reading_completed_users(
    db: AsyncSession,
    row: MagicReadingContent,
    targets: list[MagicReadingContentTarget] | None = None,
) -> int:
    user_ids = await _collect_target_user_ids(db, targets or [])
    if not user_ids:
        return 0
    result = await db.execute(
        select(func.count(func.distinct(MagicAudioUpload.user_id)))
        .where(
            MagicAudioUpload.user_id.in_(user_ids),
            MagicAudioUpload.is_deleted.is_(False),
            MagicAudioUpload.reading_content_id == row.id,
        )
    )
    return int(result.scalar_one() or 0)


async def _has_reading_checkins(
    db: AsyncSession,
    row: MagicReadingContent,
    targets: list[MagicReadingContentTarget] | None = None,
) -> bool:
    return await _count_reading_completed_users(db, row, targets) > 0


def _build_target_user_ids_from_active_users(
    active_users: list[User],
    targets: list[MagicReadingContentTarget],
) -> list[int]:
    if not targets:
        return []
    matched_ids: set[int] = set()
    for user in active_users:
        if any(_reading_target_matches_user(user, target) for target in targets):
            matched_ids.add(int(user.id))
    return sorted(matched_ids)


async def _build_reading_counts_for_rows(
    db: AsyncSession,
    rows: list[MagicReadingContent],
    targets_map: dict[int, list[MagicReadingContentTarget]],
) -> tuple[dict[int, int], dict[int, int], dict[int, bool]]:
    if not rows:
        return {}, {}, {}
    active_users = (
        await db.execute(select(User).where(User.role.in_(["user", "admin"]), User.disabled.is_(False)))
    ).scalars().all()
    target_user_map: dict[int, list[int]] = {}
    push_count_map: dict[int, int] = {}
    for row in rows:
        content_id = int(row.id)
        target_user_ids = _build_target_user_ids_from_active_users(active_users, targets_map.get(content_id, []))
        target_user_map[content_id] = target_user_ids
        push_count_map[content_id] = len(target_user_ids)
    completed_pairs_result = await db.execute(
        select(MagicAudioUpload.reading_content_id, MagicAudioUpload.user_id)
        .where(
            MagicAudioUpload.reading_content_id.in_([int(item.id) for item in rows]),
            MagicAudioUpload.is_deleted.is_(False),
        )
    )
    completed_user_map: dict[int, set[int]] = {}
    for content_id, user_id in completed_pairs_result.all():
        if user_id is None or content_id is None:
            continue
        completed_user_map.setdefault(int(content_id), set()).add(int(user_id))
    completed_count_map: dict[int, int] = {}
    lock_map: dict[int, bool] = {}
    for row in rows:
        content_id = int(row.id)
        target_user_ids = set(target_user_map.get(content_id, []))
        completed_ids = completed_user_map.get(content_id, set())
        completed_count = len(target_user_ids & completed_ids)
        completed_count_map[content_id] = completed_count
        lock_map[content_id] = completed_count > 0
    return push_count_map, completed_count_map, lock_map


def _assert_not_locked(is_locked: bool) -> None:
    if is_locked:
        raise HTTPException(
            status_code=400,
            detail="该读书内容已有打卡记录，不能修改日期、标题、图片、派发对象等核心字段，请停用后重新创建。",
        )


def _assert_locked_reading_content_update_allowed(
    row: MagicReadingContent,
    *,
    current_targets: list[MagicReadingContentTarget],
    reading_date: date,
    push_time_value: time,
    title: str,
    description: str,
    series_id: int | None,
    image_source: str,
    material_asset_id: int | None,
    image_url_text: str,
    explicit_targets: list[dict[str, str]],
    normalized_target_type: str,
    user_ids: list[int],
    department_names: list[str],
    position_names: list[str],
    job_level_names: list[str],
    employment_status_values: list[str],
    image: UploadFile | None,
    makeup_deadline_at: datetime | None,
) -> None:
    current_target_pairs = sorted(
        ((target.target_type or "").strip().lower(), (target.target_id or "").strip())
        for target in current_targets
    )
    if row.reading_date != reading_date:
        _assert_not_locked(True)
    current_push_time = row.push_time or _effective_push_at(row).time().replace(microsecond=0)
    if current_push_time != push_time_value:
        _assert_not_locked(True)
    if int(row.series_id or 0) != int(series_id or 0):
        _assert_not_locked(True)
    if (row.title or "").strip() != (title or "").strip():
        _assert_not_locked(True)
    if (row.description or "").strip() != (description or "").strip():
        _assert_not_locked(True)
    if row.makeup_deadline_at != makeup_deadline_at:
        _assert_not_locked(True)

    normalized_image_source = _normalize_image_source(image_source)
    current_image_source = _normalize_image_source(row.source_type or "upload")
    if normalized_image_source != current_image_source:
        _assert_not_locked(True)
    if normalized_image_source == "material":
        if int(row.material_asset_id or 0) != int(material_asset_id or 0):
            _assert_not_locked(True)
    elif normalized_image_source == "url":
        if (row.image_url or "").strip() != (image_url_text or "").strip():
            _assert_not_locked(True)
    else:
        if image is not None and getattr(image, "filename", ""):
            _assert_not_locked(True)

    if explicit_targets:
        next_target_pairs = sorted((item["target_type"], item["target_id"]) for item in explicit_targets)
    elif normalized_target_type == "all":
        next_target_pairs = [("all", "0")]
    elif normalized_target_type == "all_newcomers":
        next_target_pairs = [("all_newcomers", "1")]
    elif normalized_target_type == "department":
        next_target_pairs = sorted(("department", name) for name in department_names if name)
    elif normalized_target_type == "position":
        next_target_pairs = sorted(("position", name) for name in position_names if name)
    elif normalized_target_type == "job_level":
        next_target_pairs = sorted(("job_level", name) for name in job_level_names if name)
    elif normalized_target_type == "employment_status":
        next_target_pairs = sorted(("employment_status", value) for value in employment_status_values if value)
    else:
        next_target_pairs = sorted(("user", str(user_id)) for user_id in user_ids)
    if next_target_pairs != current_target_pairs:
        _assert_not_locked(True)


async def _resolve_image_payload(
    db: AsyncSession,
    *,
    admin: User,
    image_source: str,
    material_asset_id: int | None,
    image: UploadFile | None,
    image_url_text: str = "",
    embedded_image_bytes: bytes | None = None,
    embedded_image_name: str = "",
    embedded_image_mime_type: str = "",
) -> dict[str, Any]:
    normalized_image_source = _normalize_image_source(image_source)
    if normalized_image_source == "upload" and image is None and embedded_image_bytes is None:
        return {
            "source_type": "upload",
            "material_asset_id": None,
            "image_object_key": "",
            "image_url": "",
            "image_file_name": "",
            "image_mime_type": "",
            "image_size": 0,
        }
    if normalized_image_source == "material":
        if not material_asset_id:
            raise HTTPException(status_code=400, detail="请选择素材库图片。")
        material_asset = await _get_material_asset_or_403(db, material_asset_id, admin, expected_type="image")
        return {
            "source_type": "material",
            "material_asset_id": int(material_asset.id),
            "image_object_key": material_asset.object_key,
            "image_url": _build_oss_object_url(_ensure_oss_settings()["public_base_url"], material_asset.object_key),
            "image_file_name": material_asset.file_name,
            "image_mime_type": material_asset.mime_type or "image/jpeg",
            "image_size": int(material_asset.file_size or 0),
        }
    if normalized_image_source == "url":
        normalized_url = (image_url_text or "").strip()
        if not normalized_url:
            raise HTTPException(status_code=400, detail="图片 URL 不能为空。")
        return {
            "source_type": "url",
            "material_asset_id": None,
            "image_object_key": "",
            "image_url": normalized_url,
            "image_file_name": normalized_url.rsplit("/", 1)[-1][:255],
            "image_mime_type": mimetypes.guess_type(normalized_url)[0] or "image/jpeg",
            "image_size": 0,
        }
    raw: bytes
    mime_type: str
    file_name: str
    if embedded_image_bytes is not None:
        raw = embedded_image_bytes
        mime_type = (embedded_image_mime_type or "").strip() or mimetypes.guess_type(embedded_image_name)[0] or "image/png"
        file_name = embedded_image_name or "reading-content.png"
    elif image is not None:
        raw = await image.read()
        mime_type = (image.content_type or "").strip() or mimetypes.guess_type(image.filename or "")[0] or "image/jpeg"
        file_name = image.filename or "reading-content.jpg"
    else:
        raise HTTPException(status_code=400, detail="请先上传读书内容图片。")
    extension = _validate_reading_image_payload(file_name, len(raw), mime_type)
    from ._oss import _build_object_key_and_name

    object_key, stored_filename = _build_object_key_and_name(file_name or f"reading-content{extension}", extension)
    await asyncio.to_thread(_upload_binary_to_oss, object_key, raw, mime_type)
    return {
        "source_type": "upload",
        "material_asset_id": None,
        "image_object_key": object_key,
        "image_url": _build_oss_object_url(_ensure_oss_settings()["public_base_url"], object_key),
        "image_file_name": _safe_filename(file_name or stored_filename),
        "image_mime_type": mime_type,
        "image_size": len(raw),
    }


def _parse_target_names_field(raw: str) -> list[str]:
    text = (raw or "").strip()
    if not text:
        return []
    if text.startswith("["):
        values = _json_loads(text, [])
        return [str(item or "").strip() for item in values if str(item or "").strip()]
    return [item.strip() for item in text.split(",") if item.strip()]


def _owned_reading_image_cleanup_key(row: MagicReadingContent) -> str:
    if _normalize_image_source(row.source_type or "upload") != "upload":
        return ""
    if row.material_asset_id:
        return ""
    return (row.image_object_key or "").strip()


async def _create_reading_content_record(
    db: AsyncSession,
    *,
    admin: User,
    reading_date: date,
    push_time_value: time,
    title: str,
    description: str,
    image_source: str,
    material_asset_id: int | None,
    series_id: int | None,
    target_type: str,
    target_user_ids: list[int],
    target_department_ids: list[str],
    target_position_ids: list[str],
    makeup_deadline_at: datetime | None,
    target_job_level_ids: list[str] | None = None,
    target_employment_status_ids: list[str] | None = None,
    targets_payload: Any = None,
    image: UploadFile | None = None,
    image_url_text: str = "",
    embedded_image_bytes: bytes | None = None,
    embedded_image_name: str = "",
    embedded_image_mime_type: str = "",
    auto_checkin_whitelist_user_ids: set[int] | None = None,
) -> dict[str, Any]:
    normalized_title = (title or "").strip()
    if not normalized_title:
        raise HTTPException(status_code=400, detail="请输入标题。")
    explicit_targets = _normalize_explicit_reading_targets(targets_payload)
    normalized_target_type = _normalize_reading_target_type(target_type) if not explicit_targets else "mixed"
    job_level_values = list(target_job_level_ids or [])
    employment_status_values = list(target_employment_status_ids or [])
    if explicit_targets:
        valid_user_ids, valid_departments, valid_positions, push_count = [], [], [], 0
    else:
        valid_user_ids, valid_departments, valid_positions, push_count = await _validate_reading_recipients(
            db,
            target_type=normalized_target_type,
            target_user_ids=target_user_ids,
            target_department_names=target_department_ids,
            target_position_names=target_position_ids,
            target_job_level_names=job_level_values,
            target_employment_status_values=employment_status_values,
        )
    push_at = _build_push_at(reading_date, push_time_value)
    resolved_makeup_deadline = _resolve_makeup_deadline(reading_date, push_at, makeup_deadline_at)
    if resolved_makeup_deadline and resolved_makeup_deadline < push_at:
        raise HTTPException(status_code=400, detail="补卡截止时间不能早于推送时间。")
    resolved_series_id = await _validate_reading_series_id(db, admin, series_id)
    await _assert_reading_date_allowed_by_series(db, resolved_series_id, reading_date)
    image_payload = await _resolve_image_payload(
        db,
        admin=admin,
        image_source=image_source,
        material_asset_id=material_asset_id,
        image=image,
        image_url_text=image_url_text,
        embedded_image_bytes=embedded_image_bytes,
        embedded_image_name=embedded_image_name,
        embedded_image_mime_type=embedded_image_mime_type,
    )
    row = MagicReadingContent(
        series_id=resolved_series_id,
        reading_date=reading_date,
        push_time=push_time_value,
        push_at=push_at,
        makeup_deadline_at=resolved_makeup_deadline,
        title=normalized_title,
        description=(description or "").strip(),
        source_type=image_payload["source_type"],
        material_asset_id=image_payload["material_asset_id"],
        image_object_key=image_payload["image_object_key"],
        image_url=image_payload["image_url"],
        image_file_name=image_payload["image_file_name"],
        image_mime_type=image_payload["image_mime_type"],
        image_size=image_payload["image_size"],
        status=READING_CONTENT_ACTIVE,
        created_by=admin.id,
        is_deleted=False,
    )
    db.add(row)
    await db.flush()
    if explicit_targets:
        targets = await _replace_reading_targets_from_payload(db, row.id, explicit_targets)
        push_count = await _count_reading_targets(db, targets)
    else:
        targets = await _replace_reading_targets(
            db,
            row.id,
            target_type=normalized_target_type,
            user_ids=valid_user_ids,
            department_names=valid_departments,
            position_names=valid_positions,
            job_level_names=job_level_values,
            employment_status_values=employment_status_values,
        )
    await enqueue_audio_actions_for_reading_content(
        db,
        row,
        targets,
        created_by=admin.id,
        auto_checkin_whitelist_user_ids=auto_checkin_whitelist_user_ids,
    )
    await db.refresh(row)
    series = await db.get(MagicReadingSeries, resolved_series_id) if resolved_series_id else None
    return _reading_content_to_dict(
        row,
        targets=targets,
        image_url=await asyncio.to_thread(_safe_reading_image_url, row.image_object_key or "", row.image_url or ""),
        creator=admin,
        series=series,
        push_count=push_count,
    )


async def _build_reading_user_status_map(
    db: AsyncSession,
    *,
    user_id: int,
    target_date: date | None = None,
    content_ids: list[int] | None = None,
) -> dict[str, Any]:
    stmt = (
        select(MagicAudioUpload)
        .where(MagicAudioUpload.user_id == user_id, MagicAudioUpload.is_deleted.is_(False))
        .order_by(MagicAudioUpload.uploaded_on.asc())
    )
    filters = []
    if target_date:
        filters.append(MagicAudioUpload.uploaded_date == target_date)
    if content_ids:
        filters.append(MagicAudioUpload.reading_content_id.in_(content_ids))
    if filters:
        stmt = stmt.where(or_(*filters))
    result = await db.execute(stmt)
    uploads = result.scalars().all()
    by_content_id = {
        int(item.reading_content_id): item
        for item in uploads
        if item.reading_content_id
    }
    return {
        "has_upload": bool(uploads),
        "records": uploads,
        "by_content_id": by_content_id,
        "legacy_records": [item for item in uploads if not item.reading_content_id],
    }


async def _resolve_import_targets(
    db: AsyncSession,
    *,
    target_type: str,
    raw_targets: list[str],
    user_name_map: dict[str, list[User]],
    user_id_map: dict[int, User],
    department_names: set[str],
    position_names: set[str],
    job_level_names: set[str],
    employment_status_names: set[str],
) -> tuple[list[int], list[str], list[str], list[str], list[str], list[str]]:
    errors: list[str] = []
    if target_type == "all":
        return [], [], [], [], [], []
    if target_type == "all_newcomers":
        return [], [], [], [], [], []
    if target_type == "department":
        names = sorted({item for item in raw_targets if item})
        missing = [item for item in names if item not in department_names]
        if missing:
            errors.append(f"目标部门不存在：{'、'.join(missing)}")
        return [], names, [], [], [], errors
    if target_type == "position":
        names = sorted({item for item in raw_targets if item})
        missing = [item for item in names if item not in position_names]
        if missing:
            errors.append(f"目标岗位不存在：{'、'.join(missing)}")
        return [], [], names, [], [], errors
    if target_type == "job_level":
        names = sorted({item for item in raw_targets if item})
        missing = [item for item in names if item not in job_level_names]
        if missing:
            errors.append(f"职级不存在：{'、'.join(missing)}")
        return [], [], [], names, [], errors
    if target_type == "employment_status":
        names = sorted({item for item in raw_targets if item})
        missing = [item for item in names if item not in employment_status_names]
        if missing:
            errors.append(f"在职状态不存在：{'、'.join(missing)}")
        return [], [], [], [], names, errors
    user_ids: list[int] = []
    for item in raw_targets:
        if not item:
            continue
        if item.isdigit():
            user = user_id_map.get(int(item))
            if not user:
                errors.append(f"目标员工不存在：{item}")
                continue
            user_ids.append(int(user.id))
            continue
        candidates = user_name_map.get(item, [])
        if not candidates:
            errors.append(f"目标员工不存在：{item}")
            continue
        if len(candidates) > 1:
            errors.append(f"目标员工重名，请使用员工 ID：{item}")
            continue
        user_ids.append(int(candidates[0].id))
    return sorted(set(user_ids)), [], [], [], [], errors


def _extract_import_embedded_images(sheet: Any) -> dict[int, dict[str, Any]]:
    images = getattr(sheet, "_images", None) or []
    if not images:
        return {}
    result: dict[int, dict[str, Any]] = {}
    for image in images:
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        row_index = getattr(marker, "row", None)
        if row_index is None:
            continue
        excel_row = int(row_index) + 1
        try:
            raw = image._data()
        except Exception:
            continue
        if not raw:
            continue
        image_format = str(getattr(image, "format", "") or "").lower()
        if image_format == "jpg":
            image_format = "jpeg"
        mime_type = f"image/{image_format}" if image_format else "image/png"
        extension = mimetypes.guess_extension(mime_type) or ".png"
        result[excel_row] = {
            "bytes": raw,
            "mime_type": mime_type,
            "file_name": f"reading-import-row-{excel_row}{extension}",
        }
    return result


def _extract_dispimg_id(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    match = re.search(r'DISPIMG\("([^"]+)"', text, flags=re.IGNORECASE)
    return match.group(1).strip() if match else ""


def _extract_import_dispimg_images(
    workbook_bytes: bytes,
    *,
    active_sheet_title: str,
) -> dict[int, dict[str, Any]]:
    ns_main = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    ns_rel_doc = {"rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    ns_rel_pkg = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    ns_cellimg = {
        "etc": "http://www.wps.cn/officeDocument/2017/etCustomData",
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    try:
        with ZipFile(BytesIO(workbook_bytes)) as archive:
            workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
            workbook_rels_xml = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            workbook_rel_map = {
                rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
                for rel in workbook_rels_xml.findall("rel:Relationship", ns_rel_pkg)
            }
            sheet_target = ""
            for sheet in workbook_xml.findall("main:sheets/main:sheet", ns_main):
                if sheet.attrib.get("name") == active_sheet_title:
                    rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
                    sheet_target = workbook_rel_map.get(rel_id, "")
                    break
            if not sheet_target:
                return {}
            sheet_path = posixpath.normpath(posixpath.join("xl", sheet_target))
            sheet_xml = ET.fromstring(archive.read(sheet_path))

            image_name_to_row: dict[str, int] = {}
            for cell in sheet_xml.findall(".//main:c", ns_main):
                cell_ref = cell.attrib.get("r", "")
                if not cell_ref:
                    continue
                row_digits = "".join(ch for ch in cell_ref if ch.isdigit())
                if not row_digits:
                    continue
                formula_node = cell.find("main:f", ns_main)
                value_node = cell.find("main:v", ns_main)
                dispimg_id = _extract_dispimg_id(
                    formula_node.text if formula_node is not None and formula_node.text else value_node.text if value_node is not None else ""
                )
                if dispimg_id:
                    image_name_to_row[dispimg_id] = int(row_digits)

            if not image_name_to_row:
                return {}

            cellimages_xml = ET.fromstring(archive.read("xl/cellimages.xml"))
            cellimages_rels_xml = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
            rel_target_map = {
                rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
                for rel in cellimages_rels_xml.findall("rel:Relationship", ns_rel_pkg)
            }

            result: dict[int, dict[str, Any]] = {}
            for cell_image in cellimages_xml.findall("etc:cellImage", ns_cellimg):
                pic = cell_image.find("xdr:pic", ns_cellimg)
                if pic is None:
                    continue
                c_nv_pr = pic.find("xdr:nvPicPr/xdr:cNvPr", ns_cellimg)
                blip = pic.find("xdr:blipFill/a:blip", ns_cellimg)
                image_name = (c_nv_pr.attrib.get("name", "") if c_nv_pr is not None else "").strip()
                rel_id = (blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed", "") if blip is not None else "").strip()
                excel_row = image_name_to_row.get(image_name)
                media_target = rel_target_map.get(rel_id, "")
                if not excel_row or not media_target:
                    continue
                media_path = posixpath.normpath(posixpath.join("xl", media_target))
                raw = archive.read(media_path)
                mime_type = mimetypes.guess_type(media_path)[0] or "image/png"
                extension = posixpath.splitext(media_path)[1] or (mimetypes.guess_extension(mime_type) or ".png")
                result[excel_row] = {
                    "bytes": raw,
                    "mime_type": mime_type,
                    "file_name": f"reading-import-row-{excel_row}{extension}",
                }
            return result
    except Exception:
        return {}


def _normalize_series_lookup_key(value: Any) -> str:
    return str(value or "").strip().lower()


def _is_dispimg_formula(value: str) -> bool:
    text = str(value or "").strip().upper()
    return text.startswith("=DISPIMG(")


async def _resolve_import_series(
    db: AsyncSession,
    *,
    admin: User,
    series_value: Any,
    series_by_id: dict[int, MagicReadingSeries],
    series_by_title: dict[str, list[MagicReadingSeries]],
) -> tuple[int | None, list[str], list[str], str]:
    errors: list[str] = []
    warnings: list[str] = []
    text = str(series_value or "").strip()
    if not text:
        return None, errors, warnings, ""
    if text in {"未归属系列", "无", "none", "null", "-"}:
        return None, errors, warnings, "未归属系列"
    row: MagicReadingSeries | None = None
    if text.isdigit():
        row = series_by_id.get(int(text))
        if not row:
            errors.append(f"读书系列 ID 不存在：{text}")
            return None, errors, warnings, text
    else:
        candidates = series_by_title.get(_normalize_series_lookup_key(text), [])
        if not candidates:
            errors.append(f"读书系列不存在：{text}")
            return None, errors, warnings, text
        if len(candidates) > 1:
            errors.append(f"读书系列重名，请改用系列 ID：{text}")
            return None, errors, warnings, text
        row = candidates[0]
    try:
        resolved_id = await _validate_reading_series_id(db, admin, int(row.id))
    except HTTPException as exc:
        errors.append(str(exc.detail))
        return None, errors, warnings, row.title
    return resolved_id, errors, warnings, row.title


async def _resolve_import_image(
    db: AsyncSession,
    *,
    image_value: str,
    asset_by_id: dict[int, MaterialAsset],
    asset_by_name: dict[str, list[MaterialAsset]],
    embedded_image: dict[str, Any] | None = None,
) -> tuple[str, int | None, str, str, str, str, list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    text = (image_value or "").strip()
    if embedded_image and not text:
        return (
            "upload",
            None,
            "",
            base64.b64encode(embedded_image["bytes"]).decode("ascii"),
            embedded_image["file_name"],
            embedded_image["mime_type"],
            errors,
            ["已识别 Excel 内嵌图片，将按上传图片导入。"],
        )
    if _is_dispimg_formula(text):
        if embedded_image:
            warnings.append("已识别公式图片对应的 Excel 内嵌图片，将按上传图片导入。")
            return (
                "upload",
                None,
                "",
                base64.b64encode(embedded_image["bytes"]).decode("ascii"),
                embedded_image["file_name"],
                embedded_image["mime_type"],
                errors,
                warnings,
            )
        errors.append("当前不支持识别 =DISPIMG(...) 公式图片。请直接插入 Excel 图片，或填写图片 URL / 素材库资源 ID/名称。")
        return "upload", None, "", "", "", "", errors, warnings
    if not text:
        warnings.append("未识别到图片，将以空图导入。")
        return "url", None, "", "", "", "", errors, warnings
    if text.startswith("http://") or text.startswith("https://"):
        return "url", None, text, "", "", "", errors, warnings
    if text.isdigit():
        asset = asset_by_id.get(int(text))
        if not asset:
            errors.append(f"素材库资源 ID 不存在：{text}")
            return "url", None, "", "", "", "", errors, warnings
        return "material", int(asset.id), "", "", "", "", errors, warnings
    candidates = asset_by_name.get(text, [])
    if not candidates:
        if embedded_image:
            warnings.append(f"“{text}”未匹配到素材库图片，已优先使用 Excel 内嵌图片导入。")
            return (
                "upload",
                None,
                "",
                base64.b64encode(embedded_image["bytes"]).decode("ascii"),
                embedded_image["file_name"],
                embedded_image["mime_type"],
                errors,
                warnings,
            )
        errors.append(f"未匹配到素材库图片：{text}。请填写有效图片 URL、素材库资源 ID/名称，或直接插入 Excel 图片。")
        return "upload", None, "", "", "", "", errors, warnings
    if len(candidates) > 1:
        errors.append(f"素材名称重复，请改用资源 ID：{text}")
        return "url", None, "", "", "", "", errors, warnings
    return "material", int(candidates[0].id), "", "", "", "", errors, warnings


async def _parse_import_workbook(
    db: AsyncSession,
    workbook_bytes: bytes,
    *,
    admin: User,
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="未安装 openpyxl，暂时无法导入 Excel。") from exc
    workbook = load_workbook(filename=BytesIO(workbook_bytes), data_only=True)
    sheet = workbook.active
    embedded_images_by_row = _extract_import_embedded_images(sheet)
    if not embedded_images_by_row:
        embedded_images_by_row = _extract_import_dispimg_images(
            workbook_bytes,
            active_sheet_title=getattr(sheet, "title", "") or "",
        )
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 内容为空。")
    header = [str(item or "").strip() for item in rows[0]]
    expected = ["日期", "推送时间", "标题", "描述", "所属系列", "推送图片", "目标人群类型", "目标人群", "补卡截止时间"]
    if header[: len(expected)] != expected:
        raise HTTPException(status_code=400, detail="Excel 模板表头不匹配，请先下载最新模板。")

    users = (await db.execute(select(User).where(User.role.in_(["user", "admin"]), User.disabled.is_(False)))).scalars().all()
    user_id_map = {int(item.id): item for item in users}
    user_name_map: dict[str, list[User]] = {}
    for item in users:
        for key in {(item.username or "").strip(), (item.real_name or "").strip(), (item.display_name or "").strip()}:
            if key:
                user_name_map.setdefault(key, []).append(item)
    department_names = {(item.department or "").strip() for item in users if (item.department or "").strip()}
    position_names = {(item.position or "").strip() for item in users if (item.position or "").strip()}
    job_level_names = {"M线", "P线"}
    employment_status_names = {
        (item.value or "").strip()
        for item in (
            await db.execute(
                select(ConfigOption).where(ConfigOption.category == "employment_status", ConfigOption.enabled.is_(True))
            )
        ).scalars().all()
        if (item.value or "").strip()
    }
    asset_rows = (
        await db.execute(select(MaterialAsset).where(MaterialAsset.is_deleted.is_(False), MaterialAsset.asset_type == "image"))
    ).scalars().all()
    asset_by_id = {int(item.id): item for item in asset_rows}
    asset_by_name: dict[str, list[MaterialAsset]] = {}
    for item in asset_rows:
        asset_by_name.setdefault((item.name or "").strip(), []).append(item)
    series_rows = (
        await db.execute(select(MagicReadingSeries))
    ).scalars().all()
    series_by_id = {int(item.id): item for item in series_rows}
    series_by_title: dict[str, list[MagicReadingSeries]] = {}
    for item in series_rows:
        key = _normalize_series_lookup_key(item.title)
        if key:
            series_by_title.setdefault(key, []).append(item)

    parsed_rows: list[dict[str, Any]] = []
    for index, raw in enumerate(rows[1:], start=2):
        (
            date_value,
            push_time_value,
            title_value,
            description_value,
            series_value,
            image_value,
            target_type_value,
            targets_value,
            deadline_value,
        ) = (list(raw[:9]) + [None] * 9)[:9]
        embedded_image = embedded_images_by_row.get(index)
        if _is_empty_import_row(
            [
                date_value,
                push_time_value,
                title_value,
                description_value,
                series_value,
                image_value,
                target_type_value,
                targets_value,
                deadline_value,
            ],
            embedded_image,
        ):
            continue
        errors: list[str] = []
        warnings: list[str] = []
        normalized: dict[str, Any] = {}
        try:
            reading_date = _parse_date_value(date_value)
            normalized["reading_date"] = reading_date.isoformat()
        except ValueError as exc:
            errors.append(str(exc))
            reading_date = None
        try:
            push_time = _parse_time_value(push_time_value)
            normalized["push_time"] = push_time.strftime("%H:%M:%S")
        except ValueError as exc:
            errors.append(str(exc))
            push_time = None
        title_text = str(title_value or "").strip()
        if not title_text:
            errors.append("标题不能为空。")
        normalized["title"] = title_text
        normalized["description"] = str(description_value or "").strip()
        resolved_series_id, series_errors, series_warnings, series_title = await _resolve_import_series(
            db,
            admin=admin,
            series_value=series_value,
            series_by_id=series_by_id,
            series_by_title=series_by_title,
        )
        errors.extend(series_errors)
        warnings.extend(series_warnings)
        normalized["series_id"] = resolved_series_id
        normalized["series_title"] = series_title
        try:
            target_type = _normalize_import_target_type(target_type_value)
            normalized["target_type"] = target_type
        except ValueError as exc:
            errors.append(str(exc))
            target_type = "user"
            normalized["target_type"] = target_type
        raw_targets = _split_multi_text(targets_value)
        normalized["target_labels"] = raw_targets
        try:
            deadline = _parse_excel_datetime(deadline_value, field_name="补卡截止时间")
            normalized["makeup_deadline_at"] = deadline.isoformat(sep=" ") if deadline else ""
        except ValueError as exc:
            errors.append(str(exc))
            deadline = None
        (
            image_source,
            material_asset_id,
            image_url,
            embedded_image_base64,
            embedded_image_name,
            embedded_image_mime_type,
            image_errors,
            image_warnings,
        ) = await _resolve_import_image(
            db,
            image_value=str(image_value or "").strip(),
            asset_by_id=asset_by_id,
            asset_by_name=asset_by_name,
            embedded_image=embedded_image,
        )
        errors.extend(image_errors)
        warnings.extend(image_warnings)
        normalized["image_source"] = image_source
        normalized["material_asset_id"] = material_asset_id
        normalized["image_url"] = image_url
        normalized["embedded_image_base64"] = embedded_image_base64
        normalized["embedded_image_name"] = embedded_image_name
        normalized["embedded_image_mime_type"] = embedded_image_mime_type
        target_user_ids: list[int] = []
        target_department_ids: list[str] = []
        target_position_ids: list[str] = []
        target_job_level_ids: list[str] = []
        target_employment_status_ids: list[str] = []
        target_errors: list[str] = []
        if target_type not in {"all", "all_newcomers"} and not raw_targets:
            target_errors.append("目标人群不能为空。")
        else:
            (
                target_user_ids,
                target_department_ids,
                target_position_ids,
                target_job_level_ids,
                target_employment_status_ids,
                target_errors,
            ) = await _resolve_import_targets(
                db,
                target_type=target_type,
                raw_targets=raw_targets,
                user_name_map=user_name_map,
                user_id_map=user_id_map,
                department_names=department_names,
                position_names=position_names,
                job_level_names=job_level_names,
                employment_status_names=employment_status_names,
            )
        errors.extend(target_errors)
        normalized["target_user_ids"] = target_user_ids
        normalized["target_department_ids"] = target_department_ids
        normalized["target_position_ids"] = target_position_ids
        normalized["target_job_level_ids"] = target_job_level_ids
        normalized["target_employment_status_ids"] = target_employment_status_ids
        if reading_date and push_time:
            push_at = _build_push_at(reading_date, push_time)
            normalized["push_at"] = push_at.isoformat(sep=" ")
            if deadline and deadline < push_at:
                errors.append("补卡截止时间不能早于推送时间。")
            if reading_date and normalized["series_id"]:
                try:
                    await _assert_reading_date_allowed_by_series(db, normalized["series_id"], reading_date)
                except HTTPException as exc:
                    errors.append(str(exc.detail))
        parsed_rows.append(
            {
                "row_number": index,
                "raw": {
                    "日期": date_value,
                    "推送时间": push_time_value,
                    "标题": title_value,
                    "描述": description_value,
                    "所属系列": series_value,
                    "推送图片": image_value,
                    "目标人群类型": target_type_value,
                    "目标人群": targets_value,
                    "补卡截止时间": deadline_value,
                },
                "parsed": normalized,
                "errors": errors,
                "warnings": warnings,
                "can_import": not errors,
            }
        )
    return parsed_rows


@router.get("/admin/reading-series")
async def list_admin_reading_series(
    keyword: str = "",
    status: str = "",
    only_selectable: bool = False,
    page: int = 1,
    page_size: int = 20,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    stmt = select(MagicReadingSeries)
    if not is_super_admin(admin):
        stmt = stmt.where(MagicReadingSeries.created_by == admin.id)
    keyword_text = (keyword or "").strip()
    if keyword_text:
        like_value = f"%{keyword_text}%"
        stmt = stmt.where(or_(MagicReadingSeries.title.like(like_value), MagicReadingSeries.description.like(like_value)))
    status_text = (status or "").strip().lower()
    if status_text:
        if status_text == "all":
            pass
        elif status_text not in READING_SERIES_STATUSES:
            raise HTTPException(status_code=400, detail="读书系列状态不合法。")
        else:
            stmt = stmt.where(MagicReadingSeries.status == status_text)
    else:
        stmt = stmt.where(MagicReadingSeries.status != READING_SERIES_ARCHIVED)
    if only_selectable:
        stmt = stmt.where(MagicReadingSeries.status.in_(["active", "draft"]))
    page = max(int(page or 1), 1)
    page_size = max(min(int(page_size or 20), 100), 1)
    total = int((await db.execute(select(func.count()).select_from(stmt.order_by(None).subquery()))).scalar_one() or 0)
    rows = (
        await db.execute(
            stmt.order_by(desc(MagicReadingSeries.created_at), desc(MagicReadingSeries.id))
            .offset((page - 1) * page_size)
            .limit(page_size)
        )
    ).scalars().all()
    series_ids = [int(item.id) for item in rows]
    count_map: dict[int, tuple[int, int]] = {}
    out_of_range_map: dict[int, int] = {}
    targets_map = await _get_reading_series_targets_map(db, series_ids)
    if series_ids:
        content_rows = (
            await db.execute(
                select(MagicReadingContent.series_id, MagicReadingContent.reading_date, MagicReadingContent.status)
                .where(MagicReadingContent.series_id.in_(series_ids), MagicReadingContent.is_deleted.is_(False))
            )
        ).all()
        series_map_for_range = {int(item.id): item for item in rows}
        for series_id, reading_date, content_status in content_rows:
            if not series_id:
                continue
            current_total, current_enabled = count_map.get(int(series_id), (0, 0))
            count_map[int(series_id)] = (
                current_total + 1,
                current_enabled + (1 if content_status == READING_CONTENT_ACTIVE else 0),
            )
            series_row = series_map_for_range.get(int(series_id))
            if series_row and ((series_row.start_date and reading_date < series_row.start_date) or (series_row.end_date and reading_date > series_row.end_date)):
                out_of_range_map[int(series_id)] = out_of_range_map.get(int(series_id), 0) + 1
    return {
        "items": [
            _reading_series_to_dict(
                item,
                content_count=count_map.get(int(item.id), (0, 0))[0],
                enabled_content_count=count_map.get(int(item.id), (0, 0))[1],
                out_of_range_content_count=out_of_range_map.get(int(item.id), 0),
                targets=targets_map.get(int(item.id), []),
                target_summary=_reading_series_targets_summary(targets_map.get(int(item.id), [])),
            )
            for item in rows
        ],
        "page": page,
        "page_size": page_size,
        "total": total,
    }


@router.post("/admin/reading-series")
async def create_admin_reading_series(
    payload: ReadingSeriesPayload,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    if payload.start_date and payload.end_date and payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期。")
    row = MagicReadingSeries(
        title=payload.title.strip(),
        description=payload.description.strip(),
        start_date=payload.start_date,
        end_date=payload.end_date,
        status=payload.status,
        created_by=admin.id,
    )
    db.add(row)
    await db.flush()
    await _replace_reading_series_targets(db, int(row.id), payload.targets)
    await db.flush()
    await db.refresh(row)
    targets_map = await _get_reading_series_targets_map(db, [int(row.id)])
    targets = targets_map.get(int(row.id), [])
    return _reading_series_to_dict(row, targets=targets, target_summary=_reading_series_targets_summary(targets))


@router.get("/admin/reading-series/{series_id}")
async def get_admin_reading_series_detail(
    series_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_series_or_404(db, series_id)
    if not _can_manage_reading_series(admin, row):
        raise HTTPException(status_code=403, detail="无权查看该读书系列。")
    content_rows = (
        await db.execute(
            select(MagicReadingContent)
            .where(MagicReadingContent.series_id == series_id, MagicReadingContent.is_deleted.is_(False))
            .order_by(MagicReadingContent.reading_date.asc(), MagicReadingContent.push_at.asc(), MagicReadingContent.id.asc())
        )
    ).scalars().all()
    content_ids = [int(item.id) for item in content_rows]
    targets_map = await _get_reading_content_targets_map(db, content_ids)
    series_targets_map = await _get_reading_series_targets_map(db, [series_id])
    series_targets = series_targets_map.get(series_id, [])
    contents = []
    for content in content_rows:
        targets = targets_map.get(int(content.id), [])
        push_count = await _count_reading_targets(db, targets)
        completed_count = await _count_reading_completed_users(db, content, targets)
        contents.append(
            _reading_content_to_dict(
                content,
                targets=targets,
                series=row,
                push_count=push_count,
                completed_count=completed_count,
                is_locked=await _has_reading_checkins(db, content, targets),
            )
        )
        contents[-1]["out_of_range"] = _is_content_out_of_series_range(content, row)
    return _reading_series_to_dict(
        row,
        content_count=len(contents),
        enabled_content_count=sum(1 for item in contents if item.get("status") == READING_CONTENT_ACTIVE),
        out_of_range_content_count=sum(1 for item in contents if item.get("out_of_range")),
        targets=series_targets,
        target_summary=_reading_series_targets_summary(series_targets),
        contents=contents,
    )


@router.put("/admin/reading-series/{series_id}")
async def update_admin_reading_series(
    series_id: int,
    payload: ReadingSeriesPayload,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_series_or_404(db, series_id)
    if not _can_manage_reading_series(admin, row):
        raise HTTPException(status_code=403, detail="无权编辑该读书系列。")
    if payload.start_date and payload.end_date and payload.end_date < payload.start_date:
        raise HTTPException(status_code=400, detail="结束日期不能早于开始日期。")
    row.title = payload.title.strip()
    row.description = payload.description.strip()
    row.start_date = payload.start_date
    row.end_date = payload.end_date
    row.status = payload.status
    await _replace_reading_series_targets(db, int(row.id), payload.targets)
    await db.flush()
    await db.refresh(row)
    targets_map = await _get_reading_series_targets_map(db, [int(row.id)])
    targets = targets_map.get(int(row.id), [])
    return _reading_series_to_dict(row, targets=targets, target_summary=_reading_series_targets_summary(targets))


@router.post("/admin/reading-series/{series_id}/archive")
async def archive_admin_reading_series(
    series_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_series_or_404(db, series_id)
    if not _can_manage_reading_series(admin, row):
        raise HTTPException(status_code=403, detail="无权归档该读书系列。")
    row.status = READING_SERIES_ARCHIVED
    await db.flush()
    await db.refresh(row)
    content_count = int(
        (await db.execute(
            select(func.count(MagicReadingContent.id)).where(
                MagicReadingContent.series_id == series_id,
                MagicReadingContent.is_deleted.is_(False),
            )
        )).scalar_one() or 0
    )
    targets_map = await _get_reading_series_targets_map(db, [int(row.id)])
    targets = targets_map.get(int(row.id), [])
    return _reading_series_to_dict(row, content_count=content_count, targets=targets, target_summary=_reading_series_targets_summary(targets))


@router.get("/admin/reading-contents")
async def list_admin_reading_contents(
    month: str | None = None,
    date_value: str | None = Query(default=None, alias="date"),
    series_id: int | None = None,
    keyword: str = "",
    page: int = 1,
    page_size: int = 20,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    stmt = select(MagicReadingContent).where(MagicReadingContent.is_deleted.is_(False))
    if not is_super_admin(admin):
        stmt = stmt.where(MagicReadingContent.created_by == admin.id)
    if date_value:
        try:
            target_date = date.fromisoformat(date_value)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="date 格式应为 YYYY-MM-DD。") from exc
        stmt = stmt.where(MagicReadingContent.reading_date == target_date)
    elif month:
        month_start, month_end = _parse_month(month)
        stmt = stmt.where(MagicReadingContent.reading_date >= month_start, MagicReadingContent.reading_date <= month_end)
    if series_id is not None:
        if series_id > 0:
            stmt = stmt.where(MagicReadingContent.series_id == series_id)
        else:
            stmt = stmt.where(MagicReadingContent.series_id.is_(None))
    keyword_text = (keyword or "").strip()
    if keyword_text:
        like_value = f"%{keyword_text}%"
        stmt = stmt.where(or_(MagicReadingContent.title.like(like_value), MagicReadingContent.description.like(like_value)))
    page = max(int(page or 1), 1)
    page_size = max(min(int(page_size or 20), 100), 1)
    count_stmt = select(func.count()).select_from(stmt.order_by(None).subquery())
    total = int((await db.execute(count_stmt)).scalar_one() or 0)
    stmt = stmt.order_by(desc(MagicReadingContent.reading_date), desc(MagicReadingContent.push_at), desc(MagicReadingContent.created_at))
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(stmt)).scalars().all()
    content_ids = [int(item.id) for item in rows]
    targets_map = await _get_reading_content_targets_map(db, content_ids)
    series_ids = sorted({int(item.series_id) for item in rows if item.series_id})
    series_map: dict[int, MagicReadingSeries] = {}
    if series_ids:
        series_result = await db.execute(select(MagicReadingSeries).where(MagicReadingSeries.id.in_(series_ids)))
        series_map = {int(item.id): item for item in series_result.scalars().all()}
    creator_ids = sorted({int(item.created_by) for item in rows if item.created_by})
    creator_map: dict[int, User] = {}
    if creator_ids:
        creator_result = await db.execute(select(User).where(User.id.in_(creator_ids)))
        creator_map = {int(item.id): item for item in creator_result.scalars().all()}
    push_count_map, completed_count_map, lock_map = await _build_reading_counts_for_rows(db, rows, targets_map)
    items = []
    for row in rows:
        targets = targets_map.get(int(row.id), [])
        items.append(
            _reading_content_to_dict(
                row,
                targets=targets,
                image_url=await asyncio.to_thread(_safe_reading_image_url, row.image_object_key or "", row.image_url or ""),
                creator=creator_map.get(int(row.created_by)) if row.created_by else None,
                series=series_map.get(int(row.series_id)) if row.series_id else None,
                push_count=push_count_map.get(int(row.id), 0),
                is_locked=lock_map.get(int(row.id), False),
                completed_count=completed_count_map.get(int(row.id), 0),
            )
    )
    return {"items": items, "page": page, "page_size": page_size, "total": total}


# 注意：路径里包含静态字段（template / import-preview / import-confirm / batch），
# 必须放在 /{content_id} 之前，否则 FastAPI 会优先把 "template" 之类当成 content_id 解析。
@router.get("/admin/reading-contents/template")
async def download_reading_contents_template(
    admin: User = Depends(require_admin),
) -> Any:
    del admin
    return _xlsx_response(
        "reading-contents-template.xlsx",
        ["日期", "推送时间", "标题", "描述", "所属系列", "推送图片", "目标人群类型", "目标人群", "补卡截止时间"],
        [[
            "2026-06-01",
            "09:00",
            "第一章阅读",
            "请阅读第一章内容并完成打卡",
            "新员工启航计划",
            "https://example.com/reading-cover.jpg",
            "部门",
            "销售部,市场部",
            "2026-06-03 23:59",
        ], [
            "2026-06-02",
            "09:30",
            "第二章阅读",
            "图片也可以填素材库图片名称、素材 ID，或直接把图片插入到 Excel 这一行。",
            "1",
            "门店晨读封面",
            "在职状态",
            "试用,转正",
            "2026-06-04 23:59",
        ]],
    )


@router.get("/admin/reading-contents/{content_id}/image", response_model=None)
async def preview_admin_reading_content_image(
    content_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
):
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权查看该读书内容。")
    object_key = (row.image_object_key or "").strip()
    if object_key:
        signed_url = await asyncio.to_thread(_build_signed_stream_url, object_key)
        return RedirectResponse(signed_url, status_code=307)
    fallback = (row.image_url or "").strip()
    if fallback.startswith("http://") or fallback.startswith("https://"):
        return RedirectResponse(fallback, status_code=307)
    raise HTTPException(status_code=404, detail="该读书内容未配置图片。")


@router.get("/admin/reading-contents/{content_id}")
async def get_admin_reading_content_detail(
    content_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权查看该读书内容。")
    targets_map = await _get_reading_content_targets_map(db, [content_id])
    targets = targets_map.get(content_id, [])
    creator = await db.get(User, row.created_by) if row.created_by else None
    series = await db.get(MagicReadingSeries, row.series_id) if row.series_id else None
    is_locked = await _has_reading_checkins(db, row, targets)
    return _reading_content_to_dict(
        row,
        targets=targets,
        image_url=await asyncio.to_thread(_safe_reading_image_url, row.image_object_key or "", row.image_url or ""),
        creator=creator,
        series=series,
        push_count=await _count_reading_targets(db, targets),
        is_locked=is_locked,
        completed_count=await _count_reading_completed_users(db, row, targets),
    )


@router.get("/admin/reading-contents/{content_id}/push-summary")
async def get_reading_content_push_summary(
    content_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权查看该读书内容。")
    batch = await get_latest_batch(db, content_type="reading_content", content_id=content_id)
    return {"item": batch_to_dict(batch)}


@router.get("/admin/reading-contents/{content_id}/push-entries")
async def get_reading_content_push_entries(
    content_id: int,
    batch_id: int | None = Query(None, ge=1),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权查看该读书内容。")
    entries = await get_push_entries(db, content_type="reading_content", content_id=content_id, batch_id=batch_id)
    return {"items": await entries_to_dicts(db, entries)}


@router.post("/admin/reading-contents/{content_id}/push-retry")
async def retry_reading_content_push(
    content_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权操作该读书内容。")
    return await run_reading_manual_retry(db, content_id=content_id, created_by=int(admin.id))


@router.post("/admin/reading-contents/import-preview")
async def preview_reading_contents_import(
    file: UploadFile | None = File(default=None),  # CODEX_MODIFIED
    material_asset_id: int | None = Form(default=None),  # CODEX_MODIFIED
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    _cleanup_reading_import_state()
    if material_asset_id is not None:  # CODEX_MODIFIED
        material_asset = await _get_material_asset_or_403(db, material_asset_id, admin, expected_type="document")  # CODEX_MODIFIED
        if not (material_asset.file_name or "").lower().endswith(".xlsx"):  # CODEX_MODIFIED
            raise HTTPException(status_code=400, detail="仅支持导入 .xlsx 文件。")  # CODEX_MODIFIED
        bucket = _build_oss_bucket()  # CODEX_MODIFIED
        workbook_bytes = await asyncio.to_thread(lambda: bucket.get_object(material_asset.object_key).read())  # CODEX_MODIFIED
    else:  # CODEX_MODIFIED
        if file is None:  # CODEX_MODIFIED
            raise HTTPException(status_code=400, detail="请选择要导入的 .xlsx 文件。")  # CODEX_MODIFIED
        if not (file.filename or "").lower().endswith(".xlsx"):  # CODEX_MODIFIED
            raise HTTPException(status_code=400, detail="仅支持导入 .xlsx 文件。")  # CODEX_MODIFIED
        workbook_bytes = await file.read()  # CODEX_MODIFIED
    rows = await _parse_import_workbook(db, workbook_bytes, admin=admin)
    import_token = uuid4().hex
    valid_rows = []
    for item in rows:
        if not item["can_import"]:
            continue
        parsed = dict(item["parsed"])
        parsed["row_number"] = item["row_number"]
        valid_rows.append(parsed)
    _reading_import_preview_cache[import_token] = {
        "created_by": int(admin.id),
        "created_at": _now(),
        "valid_rows": valid_rows,
    }
    return {
        "import_token": import_token,
        "rows": [_serialize_preview_row_for_response(item) for item in rows],
        "summary": {
            "total": len(rows),
            "valid": sum(1 for item in rows if item["can_import"]),
            "invalid": sum(1 for item in rows if not item["can_import"]),
        },
    }


@router.post("/admin/reading-contents/import-confirm")
async def confirm_reading_contents_import(
    payload: ReadingContentImportConfirmPayload,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> ReadingContentImportStartResponse:
    del db
    _cleanup_reading_import_state()
    preview = _reading_import_preview_cache.pop(payload.import_token, None)
    if not preview or int(preview.get("created_by") or 0) != int(admin.id):
        raise HTTPException(status_code=404, detail="导入预览已失效，请重新上传 Excel。")
    valid_rows = list(preview.get("valid_rows") or [])
    if not valid_rows:
        raise HTTPException(status_code=400, detail="没有可导入的数据。")
    job_id = uuid4().hex
    job_data = {
        "job_id": job_id,
        "created_by": int(admin.id),
        "created_at": _now(),
        "updated_at": _now(),
        "status": "pending",
        "total": len(valid_rows),
        "processed": 0,
        "success_count": 0,
        "failure_count": 0,
        "error": "",
        "failures": [],
    }
    async with _reading_import_state_lock:
        _reading_import_job_cache[job_id] = job_data
    task = asyncio.create_task(_run_reading_import_job(job_id, valid_rows, int(admin.id)))
    async with _reading_import_state_lock:
        _reading_import_job_tasks[job_id] = task
    return ReadingContentImportStartResponse(job_id=job_id, status="pending", total=len(valid_rows))


@router.get("/admin/reading-contents/import-jobs/{job_id}", response_model=ReadingContentImportJobResponse)
async def get_reading_contents_import_job(
    job_id: str,
    admin: User = Depends(require_admin),
) -> ReadingContentImportJobResponse:
    _cleanup_reading_import_state()
    job = _reading_import_job_cache.get(job_id)
    if not job or int(job.get("created_by") or 0) != int(admin.id):
        raise HTTPException(status_code=404, detail="导入任务不存在或已过期。")
    return ReadingContentImportJobResponse(
        job_id=job["job_id"],
        status=str(job["status"]),
        total=int(job["total"]),
        processed=int(job["processed"]),
        success_count=int(job["success_count"]),
        failure_count=int(job["failure_count"]),
        error=str(job.get("error") or ""),
        failures=list(job.get("failures") or []),
    )


@router.post("/admin/reading-contents/batch")
async def create_admin_reading_contents_batch(
    request: Request,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    form = await request.form()
    items_json = str(form.get("items_json") or "")
    items = _json_loads(items_json, None)
    if not isinstance(items, list) or not items:
        raise HTTPException(status_code=400, detail="批量读书内容数据不能为空。")
    upload_files = {
        key: value
        for key, value in form.multi_items()
        if getattr(value, "filename", None) and callable(getattr(value, "read", None))
    }
    created_items: list[dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            raise HTTPException(status_code=400, detail="批量读书内容格式不正确。")
        client_key = str(item.get("client_key") or "").strip()
        image_file = upload_files.get(f"image_file_{client_key}") if client_key else None
        created_items.append(
            await _create_reading_content_record(
                db,
                admin=admin,
                reading_date=_parse_date_value(item.get("reading_date")),
                push_time_value=_parse_push_time_text(item.get("push_time")),
                title=str(item.get("title") or ""),
                description=str(item.get("description") or ""),
                image_source=str(item.get("image_source") or "upload"),
                material_asset_id=int(item["material_asset_id"]) if item.get("material_asset_id") else None,
                series_id=int(item["series_id"]) if item.get("series_id") else None,
                target_type=str(item.get("target_type") or "user"),
                target_user_ids=[int(v) for v in (item.get("target_user_ids") or [])],
                target_department_ids=[str(v).strip() for v in (item.get("target_department_ids") or []) if str(v).strip()],
                target_position_ids=[str(v).strip() for v in (item.get("target_position_ids") or []) if str(v).strip()],
                target_job_level_ids=[str(v).strip() for v in (item.get("target_job_level_ids") or []) if str(v).strip()],
                target_employment_status_ids=[str(v).strip() for v in (item.get("target_employment_status_ids") or []) if str(v).strip()],
                makeup_deadline_at=_parse_datetime_text(item.get("makeup_deadline_at"), field_name="补卡截止时间"),
                targets_payload=item.get("targets"),
                image=image_file,
                image_url_text=str(item.get("image_url") or ""),
            )
        )
    return {"count": len(created_items), "items": created_items}


@router.post("/admin/reading-contents")
async def create_admin_reading_content(
    reading_date: date = Form(...),
    push_time: str = Form(...),
    title: str = Form(...),
    description: str = Form(default=""),
    image_source: str = Form(default="upload"),
    material_asset_id: int | None = Form(default=None),
    series_id: int | None = Form(default=None),
    image_url: str = Form(default=""),
    target_type: str = Form(...),
    target_user_ids: str = Form(default=""),
    target_department_ids: str = Form(default=""),
    target_position_ids: str = Form(default=""),
    target_job_level_ids: str = Form(default=""),
    targets: str = Form(default=""),
    makeup_deadline_at: str = Form(default=""),
    target_employment_status_ids: str = Form(default=""),
    image: UploadFile | None = File(default=None),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    return await _create_reading_content_record(
        db,
        admin=admin,
        reading_date=reading_date,
        push_time_value=_parse_push_time_text(push_time),
        title=title,
        description=description,
        image_source=image_source,
        material_asset_id=material_asset_id,
        series_id=series_id,
        target_type=target_type,
        target_user_ids=_parse_form_id_list(target_user_ids),
        target_department_ids=_parse_target_names_field(target_department_ids),
        target_position_ids=_parse_target_names_field(target_position_ids),
        target_job_level_ids=_parse_target_names_field(target_job_level_ids),
        target_employment_status_ids=_parse_target_names_field(target_employment_status_ids),
        makeup_deadline_at=_parse_datetime_text(makeup_deadline_at, field_name="补卡截止时间"),
        targets_payload=_json_loads(targets, []) if targets else None,
        image=image,
        image_url_text=image_url,
    )


@router.post("/admin/reading-contents/batch-status")
async def update_admin_reading_contents_status_batch(
    payload: dict[str, Any] = Body(default={}),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    raw_ids = payload.get("ids") or []
    ids: list[int] = []
    for item in raw_ids:
        try:
            number = int(item)
        except (TypeError, ValueError):
            continue
        if number > 0:
            ids.append(number)
    ids = sorted(set(ids))
    if not ids:
        raise HTTPException(status_code=400, detail="请选择至少一条读书内容。")
    status = str((payload or {}).get("status") or "").strip().lower()
    if status not in {READING_CONTENT_ACTIVE, READING_STATUS_DISABLED}:
        raise HTTPException(status_code=400, detail="状态值不合法。")

    updated_ids: list[int] = []
    skipped: list[dict[str, Any]] = []
    for content_id in ids:
        row = await _get_reading_content_or_404(db, content_id)
        if not _can_manage_reading_content(admin, row):
            skipped.append({"id": content_id, "reason": "无权操作该读书内容。"})
            continue
        row.status = status
        updated_ids.append(content_id)
    await db.flush()
    return {"success": True, "updated_ids": updated_ids, "skipped": skipped}


@router.post("/admin/reading-contents/batch-delete")
async def delete_admin_reading_contents_batch(
    payload: dict[str, Any] = Body(default={}),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    raw_ids = payload.get("ids") or []
    ids: list[int] = []
    for item in raw_ids:
        try:
            number = int(item)
        except (TypeError, ValueError):
            continue
        if number > 0:
            ids.append(number)
    ids = sorted(set(ids))
    if not ids:
        raise HTTPException(status_code=400, detail="请选择至少一条读书内容。")

    deleted_ids: list[int] = []
    skipped: list[dict[str, Any]] = []
    cleanup_keys: list[str] = []
    for content_id in ids:
        row = await _get_reading_content_or_404(db, content_id)
        if not _can_manage_reading_content(admin, row):
            skipped.append({"id": content_id, "reason": "无权删除该读书内容。"})
            continue
        targets = (await _get_reading_content_targets_map(db, [content_id])).get(content_id, [])
        if await _has_reading_checkins(db, row, targets):
            skipped.append({"id": content_id, "reason": "该内容已有打卡记录，不允许删除，请使用停用。"})
            continue
        cleanup_key = _owned_reading_image_cleanup_key(row)
        row.is_deleted = True
        row.deleted_at = _now()
        if cleanup_key:
            cleanup_keys.append(cleanup_key)
        deleted_ids.append(content_id)
    await db.flush()
    await db.commit()
    schedule_oss_object_cleanup(cleanup_keys, logger=logger)
    return {"success": True, "deleted_ids": deleted_ids, "skipped": skipped}


@router.put("/admin/reading-contents/{content_id}")
async def update_admin_reading_content(
    content_id: int,
    reading_date: date = Form(...),
    push_time: str = Form(...),
    title: str = Form(...),
    description: str = Form(default=""),
    image_source: str = Form(default="upload"),
    material_asset_id: int | None = Form(default=None),
    series_id: int | None = Form(default=None),
    image_url: str = Form(default=""),
    target_type: str = Form(...),
    target_user_ids: str = Form(default=""),
    target_department_ids: str = Form(default=""),
    target_position_ids: str = Form(default=""),
    target_job_level_ids: str = Form(default=""),
    targets: str = Form(default=""),
    makeup_deadline_at: str = Form(default=""),
    target_employment_status_ids: str = Form(default=""),
    image: UploadFile | None = File(default=None),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_content_or_404(db, content_id)
    old_cleanup_key = _owned_reading_image_cleanup_key(row)
    old_image_object_key = (row.image_object_key or "").strip()
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权编辑该读书内容。")
    targets_map = await _get_reading_content_targets_map(db, [content_id])
    current_targets = targets_map.get(content_id, [])
    has_checkins = await _has_reading_checkins(db, row, current_targets)
    normalized_title = (title or "").strip()
    if not normalized_title:
        raise HTTPException(status_code=400, detail="请输入标题。")
    explicit_targets = _normalize_explicit_reading_targets(_json_loads(targets, []) if targets else None)
    normalized_target_type = _normalize_reading_target_type(target_type) if not explicit_targets else "mixed"
    job_level_values = _parse_target_names_field(target_job_level_ids)
    employment_status_values = _parse_target_names_field(target_employment_status_ids)
    if explicit_targets:
        valid_user_ids, valid_departments, valid_positions, push_count = [], [], [], 0
    else:
        valid_user_ids, valid_departments, valid_positions, push_count = await _validate_reading_recipients(
            db,
            target_type=normalized_target_type,
            target_user_ids=_parse_form_id_list(target_user_ids),
            target_department_names=_parse_target_names_field(target_department_ids),
            target_position_names=_parse_target_names_field(target_position_ids),
            target_job_level_names=job_level_values,
            target_employment_status_values=employment_status_values,
        )
    push_time_value = _parse_push_time_text(push_time)
    push_at = _build_push_at(reading_date, push_time_value)
    resolved_makeup_deadline = _resolve_makeup_deadline(
        reading_date,
        push_at,
        _parse_datetime_text(makeup_deadline_at, field_name="补卡截止时间"),
    )
    if resolved_makeup_deadline and resolved_makeup_deadline < push_at:
        raise HTTPException(status_code=400, detail="补卡截止时间不能早于推送时间。")
    resolved_series_id = await _validate_reading_series_id(db, admin, series_id)
    await _assert_reading_date_allowed_by_series(db, resolved_series_id, reading_date)
    if has_checkins:
        _assert_locked_reading_content_update_allowed(
            row,
            current_targets=current_targets,
            reading_date=reading_date,
            push_time_value=push_time_value,
            title=normalized_title,
            description=(description or "").strip(),
            series_id=resolved_series_id,
            image_source=image_source,
            material_asset_id=material_asset_id,
            image_url_text=image_url,
            explicit_targets=explicit_targets,
            normalized_target_type=normalized_target_type,
            user_ids=valid_user_ids,
            department_names=valid_departments,
            position_names=valid_positions,
            job_level_names=job_level_values,
            employment_status_values=employment_status_values,
            image=image if (image and image.filename) else None,
            makeup_deadline_at=resolved_makeup_deadline,
        )
        targets = current_targets
        push_count = await _count_reading_targets(db, targets)
    else:
        row.series_id = resolved_series_id
        row.reading_date = reading_date
        row.push_time = push_time_value
        row.push_at = push_at
        row.makeup_deadline_at = resolved_makeup_deadline
        row.title = normalized_title
        row.description = (description or "").strip()
        row.status = row.status or READING_CONTENT_ACTIVE
        if image_source in {"material", "url"} or (image is not None and image.filename):
            image_payload = await _resolve_image_payload(
                db,
                admin=admin,
                image_source=image_source,
                material_asset_id=material_asset_id,
                image=image if (image and image.filename) else None,
                image_url_text=image_url,
            )
            row.source_type = image_payload["source_type"]
            row.material_asset_id = image_payload["material_asset_id"]
            row.image_object_key = image_payload["image_object_key"]
            row.image_url = image_payload["image_url"]
            row.image_file_name = image_payload["image_file_name"]
            row.image_mime_type = image_payload["image_mime_type"]
            row.image_size = image_payload["image_size"]
        if explicit_targets:
            targets = await _replace_reading_targets_from_payload(db, row.id, explicit_targets)
            push_count = await _count_reading_targets(db, targets)
        else:
            targets = await _replace_reading_targets(
                db,
                row.id,
                target_type=normalized_target_type,
                user_ids=valid_user_ids,
                department_names=valid_departments,
                position_names=valid_positions,
                job_level_names=job_level_values,
                employment_status_values=employment_status_values,
            )
    await db.flush()
    await db.commit()
    next_cleanup_key = _owned_reading_image_cleanup_key(row)
    if old_cleanup_key and old_cleanup_key != next_cleanup_key and old_image_object_key != (row.image_object_key or "").strip():
        schedule_oss_object_cleanup([old_cleanup_key], logger=logger)
    await db.refresh(row)
    creator = await db.get(User, row.created_by) if row.created_by else None
    series = await db.get(MagicReadingSeries, row.series_id) if row.series_id else None
    has_checkins = await _has_reading_checkins(db, row, targets)
    return _reading_content_to_dict(
        row,
        targets=targets,
        image_url=await asyncio.to_thread(_safe_reading_image_url, row.image_object_key or "", row.image_url or ""),
        creator=creator,
        series=series,
        push_count=push_count,
        is_locked=has_checkins,
        completed_count=await _count_reading_completed_users(db, row, targets),
    )


@router.post("/admin/reading-contents/{content_id}/status")
async def update_admin_reading_content_status(
    content_id: int,
    payload: dict[str, Any] = Body(default={}),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权操作该读书内容。")
    status = str((payload or {}).get("status") or "").strip().lower()
    if status not in {READING_CONTENT_ACTIVE, READING_STATUS_DISABLED}:
        raise HTTPException(status_code=400, detail="状态值不合法。")
    row.status = status
    await db.flush()
    await db.refresh(row)
    targets = (await _get_reading_content_targets_map(db, [content_id])).get(content_id, [])
    creator = await db.get(User, row.created_by) if row.created_by else None
    series = await db.get(MagicReadingSeries, row.series_id) if row.series_id else None
    has_checkins = await _has_reading_checkins(db, row, targets)
    return _reading_content_to_dict(
        row,
        targets=targets,
        image_url=await asyncio.to_thread(_safe_reading_image_url, row.image_object_key or "", row.image_url or ""),
        creator=creator,
        series=series,
        push_count=await _count_reading_targets(db, targets),
        is_locked=has_checkins,
        completed_count=await _count_reading_completed_users(db, row, targets),
    )


@router.delete("/admin/reading-contents/{content_id}")
async def delete_admin_reading_content(
    content_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, bool]:
    row = await _get_reading_content_or_404(db, content_id)
    if not _can_manage_reading_content(admin, row):
        raise HTTPException(status_code=403, detail="无权删除该读书内容。")
    targets = (await _get_reading_content_targets_map(db, [content_id])).get(content_id, [])
    if await _has_reading_checkins(db, row, targets):
        raise HTTPException(status_code=400, detail="该内容已有打卡记录，不允许删除，请使用停用。")
    cleanup_key = _owned_reading_image_cleanup_key(row)
    row.is_deleted = True
    row.deleted_at = _now()
    await db.flush()
    await db.commit()
    if cleanup_key:
        schedule_oss_object_cleanup([cleanup_key], logger=logger)
    return {"success": True}


@router.get("/my/reading-contents")
async def list_my_reading_contents(
    date_value: str | None = Query(default=None, alias="date"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    target_date = date.today()
    if date_value:
        try:
            target_date = date.fromisoformat(date_value)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="date 格式应为 YYYY-MM-DD。") from exc
    now = _now()
    result = await db.execute(
        select(MagicReadingContent)
        .where(
            MagicReadingContent.is_deleted.is_(False),
            MagicReadingContent.status == READING_CONTENT_ACTIVE,
            MagicReadingContent.reading_date == target_date,
            or_(MagicReadingContent.push_at.is_(None), MagicReadingContent.push_at <= now),
        )
        .order_by(desc(MagicReadingContent.push_at), desc(MagicReadingContent.created_at), desc(MagicReadingContent.id))
    )
    rows = result.scalars().all()
    if not rows:
        return []
    targets_map = await _get_reading_content_targets_map(db, [int(item.id) for item in rows])
    upload_state = await _build_reading_user_status_map(
        db,
        user_id=int(user.id),
        target_date=target_date,
        content_ids=[int(item.id) for item in rows],
    )
    output = []
    for row in rows:
        targets = targets_map.get(int(row.id), [])
        if not targets or not any(_reading_target_matches_user(user, target) for target in targets):
            continue
        upload = upload_state["by_content_id"].get(int(row.id))
        completed = bool(upload)
        status = "已完成" if completed else "已过补卡时间" if row.makeup_deadline_at and now > row.makeup_deadline_at else "未完成"
        can_submit = True
        submit_disabled_reason = ""
        if now < _effective_push_at(row):
            can_submit = False
            submit_disabled_reason = "未到推送时间"
        elif row.makeup_deadline_at and now > row.makeup_deadline_at:
            can_submit = False
            submit_disabled_reason = "已超过补卡截止时间，无法补交。"
        elif completed:
            can_submit = False
            submit_disabled_reason = "该读书内容已完成。"
        output.append(
            _reading_content_to_dict(
                row,
                targets=targets,
                image_url=await asyncio.to_thread(_safe_reading_image_url, row.image_object_key or "", row.image_url or ""),
                current_status=status,
                completed=completed,
                upload_id=int(upload.id) if upload else None,
                can_submit=can_submit,
                submit_disabled_reason=submit_disabled_reason,
            )
        )
    return output
