"""管理员用户 CRUD + 分页搜索 + 批量导入 + 模板下载。"""
from __future__ import annotations

import io
import json
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import delete as sql_delete, func, or_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from .access import is_super_admin
from .auth import md5_password, require_admin
from .db import get_db
from .employee_open_client import EmployeeOpenApiError
from .employee_open_client import EmployeeOpenClient
from .employee_sync import (
    build_employee_sync_preview_with_token,
    consume_preview,
    execute_employee_sync,
)
from .models import User, WecomSyncBatch, WecomSyncEntry

router = APIRouter(prefix="/api/admin/users", tags=["admin-users"])


class UserDTO(BaseModel):
    id: int
    username: str
    display_name: str
    real_name: str
    department: str
    position: str
    job_level: str = "M线"
    rank_name: str = ""
    role: str
    is_newcomer: bool
    employment_status: str = ""
    status: str
    disabled: bool
    wecom_userid: str = ""
    wecom_synced_at: str | None = None
    sync_issue_action: str = ""
    sync_issue_reason: str = ""
    created_at: str = ""
    updated_at: str = ""


class UserCreateRequest(BaseModel):
    username: str = Field(..., min_length=1, max_length=64)
    password: str = Field(..., min_length=1, max_length=256)
    display_name: str = Field(default="", max_length=128)
    real_name: str = Field(default="", max_length=128)
    department: str = Field(default="", max_length=128)
    position: str = Field(default="", max_length=128)
    job_level: str = Field(default="M线", max_length=16)
    rank_name: str = Field(default="", max_length=32)
    role: str = Field(default="user", max_length=16)
    is_newcomer: bool = False
    employment_status: str = Field(default="", max_length=32)
    status: str = Field(default="active", max_length=16)
    disabled: bool = False

    @field_validator("username", "display_name", "real_name", "department", "position", "job_level", "rank_name", "employment_status", mode="before")
    @classmethod
    def _strip(cls, v: str) -> str:
        return (v or "").strip()

    @field_validator("job_level")
    @classmethod
    def _job_level(cls, v: str) -> str:
        text = (v or "M线").strip().upper()
        if text in {"M", "M线", "M-LINE", "MLINE"}:
            return "M线"
        if text in {"P", "P线", "P-LINE", "PLINE"}:
            return "P线"
        return "M线"

    @field_validator("role")
    @classmethod
    def _role(cls, v: str) -> str:
        v = (v or "user").strip().lower()
        return v if v in ("super_admin", "admin", "user") else "user"

    @field_validator("status")
    @classmethod
    def _status(cls, v: str) -> str:
        v = (v or "active").strip().lower()
        return v if v in ("active", "inactive") else "active"


class UserUpdateRequest(BaseModel):
    """每个字段都是可选；None 表示不改。password 单独一字段 → 重置密码。"""
    password: str | None = Field(default=None, max_length=256)
    display_name: str | None = Field(default=None, max_length=128)
    real_name: str | None = Field(default=None, max_length=128)
    department: str | None = Field(default=None, max_length=128)
    position: str | None = Field(default=None, max_length=128)
    job_level: str | None = Field(default=None, max_length=16)
    rank_name: str | None = Field(default=None, max_length=32)
    role: str | None = None
    is_newcomer: bool | None = None
    employment_status: str | None = Field(default=None, max_length=32)
    status: str | None = Field(default=None, max_length=16)
    disabled: bool | None = None

    @field_validator("role")
    @classmethod
    def _role(cls, v: str | None) -> str | None:
        if v is None:
            return None
        v = v.strip().lower()
        return v if v in ("super_admin", "admin", "user") else "user"

    @field_validator("job_level")
    @classmethod
    def _job_level(cls, v: str | None) -> str | None:
        if v is None:
            return None
        text = (v or "M线").strip().upper()
        if text in {"M", "M线", "M-LINE", "MLINE"}:
            return "M线"
        if text in {"P", "P线", "P-LINE", "PLINE"}:
            return "P线"
        return "M线"

    @field_validator("status")
    @classmethod
    def _status(cls, v: str | None) -> str | None:
        if v is None:
            return None
        v = v.strip().lower()
        return v if v in ("active", "inactive") else "active"


class UserPageResponse(BaseModel):
    items: list[UserDTO]
    total: int
    page: int
    page_size: int


class BulkImportSummary(BaseModel):
    total: int
    created: int
    skipped: int
    failed: int
    errors: list[str] = Field(default_factory=list)


class BulkUserIdsPayload(BaseModel):
    ids: list[int] = Field(..., min_length=1)


class EmployeeSyncPreviewRequest(BaseModel):
    initial_mode: bool = True


class EmployeeSyncItemDTO(BaseModel):
    action: str
    reason: str = ""
    match_type: str = ""
    local_user_id: int | None = None
    local_username: str = ""
    local_name: str = ""
    local_role: str = ""
    external_user_id: int | None = None
    source_hr_status: int | None = None
    source_employment_status: int | None = None
    source_name: str = ""
    wecom_name: str = ""
    wecom_userid: str = ""
    mobile: str = ""
    department: str = ""
    position: str = ""
    rank_name: str = ""
    job_level: str = ""
    local_snapshot: dict[str, Any] | None = None


class EmployeeSyncPreviewResponse(BaseModel):
    initial_mode: bool
    total_source_users: int
    summary: dict[str, int]
    items: list[EmployeeSyncItemDTO]
    preview_token: str = ""


class EmployeeSyncExecuteRequest(BaseModel):
    initial_mode: bool = True
    preview_token: str | None = None


class EmployeeSyncExecuteResponse(BaseModel):
    batch_id: int
    initial_mode: bool
    summary: dict[str, int]


class EmployeeSearchResponse(BaseModel):
    total: int
    items: list[dict[str, object]]


class SyncBatchDTO(BaseModel):
    id: int
    mode: str
    initial_mode: bool
    total_wecom_users: int
    matched_count: int
    bound_count: int
    updated_count: int
    created_count: int
    left_count: int
    disabled_count: int
    conflict_count: int
    skipped_count: int
    executed_by: int | None = None
    executed_by_name: str = ""
    started_at: str = ""
    finished_at: str | None = None


class SyncBatchPageResponse(BaseModel):
    items: list[SyncBatchDTO]
    total: int
    page: int
    page_size: int


class SyncEntryDTO(BaseModel):
    id: int
    user_id: int | None = None
    wecom_userid: str = ""
    mobile: str = ""
    match_type: str = ""
    action: str
    status: str
    reason: str = ""
    before: dict[str, Any] | None = None
    after: dict[str, Any] | None = None
    created_at: str = ""


def _parse_snapshot(raw: str | None) -> dict[str, Any] | None:
    if not raw:
        return None
    try:
        value = json.loads(raw)
    except (TypeError, ValueError):
        return None
    return value if isinstance(value, dict) else None


def _digest(raw: str) -> str:
    raw = (raw or "").strip()
    if len(raw) == 32 and all(c in "0123456789abcdefABCDEF" for c in raw):
        return raw.lower()
    return md5_password(raw)


def _to_dto(user: User) -> UserDTO:
    return UserDTO(
        id=user.id,
        username=user.username,
        display_name=user.display_name or "",
        real_name=user.real_name or "",
        department=user.department or "",
        position=user.position or "",
        job_level=user.job_level or "M线",
        rank_name=user.rank_name or "",
        role=user.role or "user",
        is_newcomer=bool(user.is_newcomer),
        employment_status=user.employment_status or "",
        status=user.status or "active",
        disabled=bool(user.disabled),
        wecom_userid=user.wecom_userid or "",
        wecom_synced_at=user.wecom_synced_at.isoformat() if user.wecom_synced_at else None,
        sync_issue_action="sync_ok",
        sync_issue_reason="最近一次同步成功",
        created_at=user.created_at.isoformat() if user.created_at else "",
        updated_at=user.updated_at.isoformat() if user.updated_at else "",
    )


async def _attach_sync_issue_info(db: AsyncSession, users: list[UserDTO]) -> list[UserDTO]:
    if not users:
        return users
    user_ids = [int(user.id) for user in users]
    entries = (
        await db.execute(
            select(WecomSyncEntry)
            .where(WecomSyncEntry.user_id.in_(user_ids))
            .order_by(WecomSyncEntry.user_id.asc(), WecomSyncEntry.id.desc())
        )
    ).scalars().all()
    latest_by_user_id: dict[int, WecomSyncEntry] = {}
    for entry in entries:
        if entry.user_id is None:
            continue
        key = int(entry.user_id)
        if key not in latest_by_user_id:
            latest_by_user_id[key] = entry
    for user in users:
        latest = latest_by_user_id.get(int(user.id))
        if latest is None:
            continue
        latest_status = (latest.status or "").strip()
        latest_action = (latest.action or "").strip()
        if latest_action == "mark_left":
            user.sync_issue_action = "mark_left"
            user.sync_issue_reason = latest.reason or "该账号已在同步中置为离职并禁用。"
            continue
        if latest_status in {"applied", "created"}:
            continue
        user.sync_issue_action = latest_action or "sync_ok"
        user.sync_issue_reason = latest.reason or ""
    return users


def _is_manageable_by(actor: User, target: User) -> bool:
    if is_super_admin(actor):
        return True
    return not is_super_admin(target)


def _assert_manageable(actor: User, target: User) -> None:
    if not _is_manageable_by(actor, target):
        raise HTTPException(status_code=403, detail="无权操作超级管理员账号。")


# ---------- 列表 / 分页搜索 ----------


@router.get("", response_model=list[UserDTO])
async def list_users(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> list[UserDTO]:
    """派发面板用的全量列表：默认排除离职 / 禁用员工。

    这些员工已不参与任何业务（派发、推送、提醒），不应出现在选人弹窗里。
    需要查全量（含离职）的入口请走 /users/search?include_disabled=true。
    """
    stmt = select(User).where(User.disabled.is_(False)).order_by(User.id.asc())
    if not is_super_admin(admin):
        stmt = stmt.where(User.role != "super_admin")
    result = await db.execute(stmt)
    return await _attach_sync_issue_info(db, [_to_dto(u) for u in result.scalars().all()])


@router.get("/search", response_model=UserPageResponse)
async def search_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=200),
    keyword: str | None = Query(None, description="按用户名 / 真实姓名 / 显示名模糊搜索"),
    department: str | None = Query(None, description="按部门精确筛选"),
    role: str | None = Query(None, description="可选：admin / user"),
    employment_status: str | None = Query(None, description="按在职状态精确筛选"),
    include_disabled: bool = Query(
        True,
        description="是否包含已禁用 / 离职员工。用户管理页保持 True，派发选人弹窗传 False。",
    ),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> UserPageResponse:
    stmt = select(User)
    count_stmt = select(func.count()).select_from(User)
    if not is_super_admin(admin):
        stmt = stmt.where(User.role != "super_admin")
        count_stmt = count_stmt.where(User.role != "super_admin")

    if not include_disabled:
        # 派发 / 选人场景：离职（disabled=True）的员工不参与任何业务，不出现在选项里。
        stmt = stmt.where(User.disabled.is_(False))
        count_stmt = count_stmt.where(User.disabled.is_(False))

    if keyword:
        kw = f"%{keyword.strip()}%"
        cond = or_(User.username.like(kw), User.real_name.like(kw), User.display_name.like(kw))
        stmt = stmt.where(cond)
        count_stmt = count_stmt.where(cond)
    if department:
        stmt = stmt.where(User.department == department.strip())
        count_stmt = count_stmt.where(User.department == department.strip())
    if role and role.lower() in ("admin", "user"):
        stmt = stmt.where(User.role == role.lower())
        count_stmt = count_stmt.where(User.role == role.lower())
    if employment_status:
        stmt = stmt.where(User.employment_status == employment_status.strip())
        count_stmt = count_stmt.where(User.employment_status == employment_status.strip())

    total = (await db.execute(count_stmt)).scalar_one()
    stmt = stmt.order_by(User.id.desc()).limit(page_size).offset((page - 1) * page_size)
    rows = (await db.execute(stmt)).scalars().all()
    items = await _attach_sync_issue_info(db, [_to_dto(u) for u in rows])
    return UserPageResponse(
        items=items,
        total=int(total),
        page=page,
        page_size=page_size,
    )


@router.get("/departments", response_model=list[str])
async def list_departments(
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> list[str]:
    del admin
    rows = (
        await db.execute(
            select(User.department)
            .where(User.department != "")
            .group_by(User.department)
            .order_by(User.department.asc())
        )
    ).scalars().all()
    return [d for d in rows if d]


# ---------- 模板下载 / 批量导入 ----------
# 注意：路径里包含静态字段（template / bulk-import），必须放在 /{user_id} 之前，
# 否则 FastAPI 会按声明顺序优先匹配 /{user_id}，把 "template" 当成 user_id 解析。


_USER_TEMPLATE_HEADERS = [
    "用户名*", "密码*", "真实姓名", "显示名",
    "部门", "岗位", "职级", "角色", "是否新人", "在职状态",
]


def _build_user_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用户导入模板"

    header_fill = PatternFill(start_color="FFE6F0FA", end_color="FFE6F0FA", fill_type="solid")
    required_fill = PatternFill(start_color="FFFFF1F0", end_color="FFFFF1F0", fill_type="solid")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(_USER_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.fill = required_fill if header.endswith("*") else header_fill
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    samples = [
        ["zhangsan", "123456", "张三", "三哥", "销售一部", "招商主管", "M线", "普通用户", "否", "转正"],
        ["lisi",     "123456", "李四", "",     "销售一部", "招商专员", "P线", "普通用户", "是", "试岗"],
        ["wangwu",   "abcd1234", "王五", "",   "运营部",   "运营经理", "M线", "管理员",   "否", "试用"],
    ]
    for r, row in enumerate(samples, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)

    notes = [
        "",
        "填写说明：",
        "1) 用户名 / 密码必填，用户名重复将自动跳过。",
        "2) 职级：M线 / P线，默认 M线。",
        "3) 角色：普通用户 / 管理员（也可写英文 user / admin），默认普通用户。",
        "4) 是否新人：是 / 否（默认 否）。",
        "5) 在职状态可填写：试岗 / 试用 / 转正 / 离职；为空则不设置。",
        "6) 部门 / 岗位 / 显示名 可空；显示名为空时会取真实姓名或用户名。",
        "7) 密码以明文写入，后端会做 md5 后入库。",
    ]
    for i, line in enumerate(notes, start=len(samples) + 3):
        cell = ws.cell(row=i, column=1, value=line)
        ws.merge_cells(start_row=i, end_row=i, start_column=1, end_column=len(_USER_TEMPLATE_HEADERS))
        if line.startswith("填写说明"):
            cell.font = Font(bold=True, color="FF1677FF")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/template")
async def download_user_template(
    admin: User = Depends(require_admin),
):
    del admin
    data = _build_user_template()

    def _iter():
        yield data

    headers = {"Content-Disposition": 'attachment; filename="users_template.xlsx"'}
    return StreamingResponse(
        _iter(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


_ROLE_ALIASES = {
    "管理员": "admin", "admin": "admin", "administrator": "admin",
    "普通用户": "user", "user": "user", "员工": "user", "学员": "user",
}
_BOOL_TRUE = {"是", "y", "yes", "true", "1", "新人"}
_EMPLOYMENT_STATUS_ALIASES = {
    "试岗": "试岗",
    "试岗期": "试岗",
    "试用": "试用",
    "试用期": "试用",
    "转正": "转正",
    "正式": "转正",
    "在职": "转正",
    "离职": "离职",
    "禁用": "离职",
}
_JOB_LEVEL_ALIASES = {
    "": "M线",
    "m": "M线",
    "m线": "M线",
    "m-line": "M线",
    "mline": "M线",
    "p": "P线",
    "p线": "P线",
    "p-line": "P线",
    "pline": "P线",
}


def _norm_role(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    return _ROLE_ALIASES.get(text, "user")


def _norm_bool(raw: Any) -> bool:
    text = str(raw or "").strip().lower()
    return text in _BOOL_TRUE


def _norm_employment_status(raw: Any) -> str:
    text = str(raw or "").strip()
    return _EMPLOYMENT_STATUS_ALIASES.get(text, text)


def _norm_job_level(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    return _JOB_LEVEL_ALIASES.get(text, "M线")


def _line_from_rank(rank: Any) -> str:
    """原始职级（如 M3 / P0 / L1）→ 派发用的线。空/未知默认 M线。"""
    text = str(rank or "").strip().upper()
    if text:
        head = text[0]
        if head == "P":
            return "P线"
        if head == "L":
            return "L线"
    return "M线"


def _is_left_employment_status(value: str | None) -> bool:
    return (value or "").strip() == "离职"


async def _read_upload_with_limit(file: UploadFile, *, limit: int = 20 * 1024 * 1024) -> bytes:
    try:
        file.file.seek(0, 2)
        size = file.file.tell()
        file.file.seek(0)
        if size > limit:
            raise HTTPException(status_code=413, detail="上传文件不能超过 20MB。")
    except HTTPException:
        raise
    except (AttributeError, OSError):
        await file.seek(0)
    content = await file.read(limit + 1)
    if len(content) > limit:
        raise HTTPException(status_code=413, detail="上传文件不能超过 20MB。")
    return content


@router.post("/bulk-import", response_model=BulkImportSummary)
async def bulk_import_users(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> BulkImportSummary:
    """上传 Excel (.xlsx) 一次性导入多个用户。

    - 用户名重复 → skipped（不覆盖）
    - 必填字段缺失 / 数据非法 → failed，错误信息记录在 errors[]
    - 其余有效行 → created
    """
    del admin
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 格式。")
    content = await _read_upload_with_limit(file)

    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"无法解析 Excel：{exc}") from exc

    if not rows or len(rows) < 2:
        raise HTTPException(status_code=400, detail="表格为空或没有数据行。")

    data_rows = rows[1:]
    summary = BulkImportSummary(total=0, created=0, skipped=0, failed=0)

    existing = set(
        (await db.execute(select(User.username))).scalars().all()
    )

    for idx, row in enumerate(data_rows):
        if not row or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        summary.total += 1
        cells = list(row) + [None] * max(0, len(_USER_TEMPLATE_HEADERS) - len(row))
        username, password, real_name, display_name, department, position = cells[:6]
        if len(row) >= 10:
            job_level_raw, role_raw, newcomer_raw, employment_status = cells[6:10]
        else:
            job_level_raw = "M线"
            role_raw, newcomer_raw, employment_status = cells[6:9]

        username = str(username or "").strip()
        password = str(password or "").strip()
        line_no = idx + 2

        if not username:
            summary.failed += 1
            summary.errors.append(f"第 {line_no} 行：用户名不能为空")
            continue
        if not password:
            summary.failed += 1
            summary.errors.append(f"第 {line_no} 行（{username}）：密码不能为空")
            continue
        if username in existing:
            summary.skipped += 1
            continue

        real_name = str(real_name or "").strip()
        display_name = str(display_name or "").strip() or real_name or username
        department = str(department or "").strip()
        position = str(position or "").strip()
        job_level = _norm_job_level(job_level_raw)
        role = _norm_role(role_raw)
        is_newcomer = _norm_bool(newcomer_raw)
        employment_status = _norm_employment_status(employment_status)

        # 每行用 savepoint 保护：单行 IntegrityError 只回滚这一行，
        # 之前 add 但还未 commit 的行仍留在主事务里，最后一次 commit 落盘。
        try:
            async with db.begin_nested():
                user = User(
                    username=username,
                    password_md5=_digest(password),
                    display_name=display_name,
                    real_name=real_name or display_name,
                    department=department,
                    position=position,
                    job_level=job_level,
                    role=role,
                    is_newcomer=is_newcomer,
                    employment_status=employment_status,
                    status="active",
                    disabled=False,
                )
                db.add(user)
                await db.flush()
            existing.add(username)
            summary.created += 1
        except IntegrityError:
            summary.failed += 1
            summary.errors.append(f"第 {line_no} 行（{username}）：用户名冲突或数据库异常")

    return summary


# ---------- 批量操作 ----------


@router.post("/bulk-delete")
async def bulk_delete_users(
    payload: BulkUserIdsPayload,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict[str, Any]:
    rows = (
        await db.execute(select(User).where(User.id.in_(payload.ids)))
    ).scalars().all()

    deletable_ids: list[int] = []
    skipped = 0

    for user in rows:
        if user.id == admin.id or not _is_manageable_by(admin, user):
            skipped += 1
            continue
        deletable_ids.append(user.id)

    if not deletable_ids:
        return {"success": True, "deleted": 0, "skipped": skipped}

    res = await db.execute(sql_delete(User).where(User.id.in_(deletable_ids)))
    return {
        "success": True,
        "deleted": int(res.rowcount or 0),
        "skipped": skipped,
    }


@router.post("/employee-sync/preview", response_model=EmployeeSyncPreviewResponse)
async def preview_employee_sync(
    payload: EmployeeSyncPreviewRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> EmployeeSyncPreviewResponse:
    del admin
    try:
        preview, preview_token = await build_employee_sync_preview_with_token(
            db, initial_mode=payload.initial_mode
        )
    except EmployeeOpenApiError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    items = [
        EmployeeSyncItemDTO(
            action=item["action"],
            reason=item.get("reason") or "",
            match_type=item.get("match_type") or "",
            local_user_id=item.get("local_user_id"),
            local_username=item.get("local_username") or "",
            local_name=item.get("local_name") or "",
            local_role=item.get("local_role") or "",
            external_user_id=item.get("external_user_id"),
            source_hr_status=item.get("source_hr_status"),
            source_employment_status=item.get("source_employment_status"),
            source_name=item.get("source_name") or "",
            wecom_name=item.get("wecom_name") or item.get("source_name") or "",
            wecom_userid=item.get("wecom_userid") or "",
            mobile=item.get("mobile") or "",
            department=item.get("department") or "",
            position=item.get("position") or "",
            rank_name=item.get("rank_name") or "",
            job_level=item.get("job_level") or "",
            local_snapshot=item.get("local_snapshot"),
        )
        for item in preview.get("items", [])
    ]
    return EmployeeSyncPreviewResponse(
        initial_mode=bool(preview.get("initial_mode")),
        total_source_users=int(preview.get("total_source_users") or 0),
        summary={str(k): int(v) for k, v in (preview.get("summary") or {}).items()},
        items=items,
        preview_token=preview_token,
    )


@router.post("/employee-sync/execute", response_model=EmployeeSyncExecuteResponse)
async def run_employee_sync(
    payload: EmployeeSyncExecuteRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> EmployeeSyncExecuteResponse:
    cached_preview = consume_preview(payload.preview_token or "")
    try:
        result = await execute_employee_sync(
            db,
            actor=admin,
            initial_mode=payload.initial_mode,
            preview=cached_preview,
        )
    except EmployeeOpenApiError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return EmployeeSyncExecuteResponse(
        batch_id=int(result["batch_id"]),
        initial_mode=bool(result["initial_mode"]),
        summary={str(k): int(v) for k, v in (result.get("summary") or {}).items()},
    )


@router.get("/employee-sync/search", response_model=EmployeeSearchResponse)
async def search_external_employees(
    name: str | None = None,
    mobile: str | None = None,
    external_user_id: str | None = None,
    admin: User = Depends(require_admin),
) -> EmployeeSearchResponse:
    del admin
    filters: dict[str, object] = {}
    if name and name.strip():
        filters["name"] = name.strip()
    if mobile and mobile.strip():
        filters["mobile"] = mobile.strip()
    if external_user_id and external_user_id.strip():
        filters["external_user_id"] = external_user_id.strip()
    try:
        items = await EmployeeOpenClient().fetch_employees(filters)
    except EmployeeOpenApiError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return EmployeeSearchResponse(total=len(items), items=items)


@router.get("/employee-sync/batches", response_model=SyncBatchPageResponse)
async def list_sync_batches(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> SyncBatchPageResponse:
    del admin
    total = (
        await db.execute(select(func.count()).select_from(WecomSyncBatch))
    ).scalar_one()
    rows = (
        await db.execute(
            select(WecomSyncBatch)
            .order_by(WecomSyncBatch.id.desc())
            .limit(page_size)
            .offset((page - 1) * page_size)
        )
    ).scalars().all()

    executor_ids = {int(b.executed_by) for b in rows if b.executed_by}
    name_by_id: dict[int, str] = {}
    if executor_ids:
        executors = (
            await db.execute(select(User).where(User.id.in_(executor_ids)))
        ).scalars().all()
        for user in executors:
            name_by_id[int(user.id)] = (
                user.real_name or user.display_name or user.username or ""
            )

    items = [
        SyncBatchDTO(
            id=int(b.id),
            mode=b.mode or "",
            initial_mode=bool(b.initial_mode),
            total_wecom_users=int(b.total_wecom_users or 0),
            matched_count=int(b.matched_count or 0),
            bound_count=int(b.bound_count or 0),
            updated_count=int(b.updated_count or 0),
            created_count=int(b.created_count or 0),
            left_count=int(b.left_count or 0),
            disabled_count=int(b.disabled_count or 0),
            conflict_count=int(b.conflict_count or 0),
            skipped_count=int(b.skipped_count or 0),
            executed_by=int(b.executed_by) if b.executed_by else None,
            executed_by_name=name_by_id.get(int(b.executed_by), "") if b.executed_by else "",
            started_at=b.started_at.isoformat() if b.started_at else "",
            finished_at=b.finished_at.isoformat() if b.finished_at else None,
        )
        for b in rows
    ]
    return SyncBatchPageResponse(
        items=items,
        total=int(total),
        page=page,
        page_size=page_size,
    )


@router.get("/employee-sync/batches/{batch_id}/entries", response_model=list[SyncEntryDTO])
async def list_sync_batch_entries(
    batch_id: int,
    action: str | None = Query(None, description="可选：按 action 过滤"),
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> list[SyncEntryDTO]:
    del admin
    stmt = select(WecomSyncEntry).where(WecomSyncEntry.batch_id == batch_id)
    if action and action.strip():
        stmt = stmt.where(WecomSyncEntry.action == action.strip())
    rows = (await db.execute(stmt.order_by(WecomSyncEntry.id.asc()))).scalars().all()
    return [
        SyncEntryDTO(
            id=int(e.id),
            user_id=int(e.user_id) if e.user_id else None,
            wecom_userid=e.wecom_userid or "",
            mobile=e.mobile or "",
            match_type=e.match_type or "",
            action=e.action or "",
            status=e.status or "",
            reason=e.reason or "",
            before=_parse_snapshot(e.before_json),
            after=_parse_snapshot(e.after_json),
            created_at=e.created_at.isoformat() if e.created_at else "",
        )
        for e in rows
    ]


# ---------- 单条 CRUD ----------


@router.get("/{user_id}", response_model=UserDTO)
async def get_user_detail(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> UserDTO:
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在。")
    _assert_manageable(admin, user)
    return _to_dto(user)


@router.post("", response_model=UserDTO)
async def create_user(
    payload: UserCreateRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> UserDTO:
    if payload.role == "super_admin" and not is_super_admin(admin):
        raise HTTPException(status_code=403, detail="仅超级管理员可创建超级管理员账号。")
    employment_status = payload.employment_status or ""
    disabled = bool(payload.disabled) or _is_left_employment_status(employment_status)
    user = User(
        username=payload.username,
        password_md5=_digest(payload.password),
        display_name=payload.display_name or payload.real_name or payload.username,
        real_name=payload.real_name or payload.display_name or payload.username,
        department=payload.department or "",
        position=payload.position or "",
        job_level=_line_from_rank(payload.rank_name),
        rank_name=payload.rank_name or "",
        role=payload.role,
        is_newcomer=payload.is_newcomer,
        employment_status=employment_status,
        status="inactive" if disabled else payload.status,
        disabled=disabled,
    )
    db.add(user)
    try:
        await db.flush()
    except IntegrityError as exc:
        raise HTTPException(status_code=409, detail="用户名已存在。") from exc
    await db.refresh(user)
    return _to_dto(user)


@router.put("/{user_id}", response_model=UserDTO)
async def update_user(
    user_id: int,
    payload: UserUpdateRequest,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> UserDTO:
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在。")
    _assert_manageable(admin, user)
    if payload.password is not None and payload.password.strip():
        user.password_md5 = _digest(payload.password)
    if payload.display_name is not None:
        user.display_name = payload.display_name.strip()
    if payload.real_name is not None:
        user.real_name = payload.real_name.strip()
    if payload.department is not None:
        user.department = payload.department.strip()
    if payload.position is not None:
        user.position = payload.position.strip()
    if payload.rank_name is not None:
        user.rank_name = payload.rank_name.strip()
        user.job_level = _line_from_rank(user.rank_name)
    elif payload.job_level is not None:
        user.job_level = payload.job_level or "M线"
    if payload.role is not None:
        if payload.role == "super_admin" and not is_super_admin(admin):
            raise HTTPException(status_code=403, detail="仅超级管理员可设置超级管理员角色。")
        if user.id == admin.id:
            expected_self_role = "super_admin" if is_super_admin(admin) else "admin"
            if payload.role != expected_self_role:
                raise HTTPException(status_code=400, detail="不能降级当前登录的管理员。")
        user.role = payload.role
    if payload.is_newcomer is not None:
        user.is_newcomer = payload.is_newcomer
    if payload.employment_status is not None:
        user.employment_status = payload.employment_status.strip()
    if payload.status is not None:
        user.status = payload.status
    if payload.disabled is not None:
        if user.id == admin.id and payload.disabled:
            raise HTTPException(status_code=400, detail="不能禁用当前登录的管理员。")
        user.disabled = payload.disabled
    if _is_left_employment_status(user.employment_status):
        if user.id == admin.id:
            raise HTTPException(status_code=400, detail="不能将当前登录的管理员置为离职。")
        user.disabled = True
        user.status = "inactive"
    await db.flush()
    await db.refresh(user)
    return _to_dto(user)


@router.delete("/{user_id}")
async def delete_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    admin: User = Depends(require_admin),
) -> dict:
    if user_id == admin.id:
        raise HTTPException(status_code=400, detail="不能删除当前登录的管理员。")
    target = await db.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="用户不存在。")
    _assert_manageable(admin, target)
    res = await db.execute(sql_delete(User).where(User.id == user_id))
    if res.rowcount == 0:
        raise HTTPException(status_code=404, detail="用户不存在。")
    return {"success": True}
